"""
Сборка данных бухгалтерского отчёта за месяц + запись в Excel.

Переиспользует существующие сервисы финансов — не дублирует правила
half-lesson и FIFO-остаток аванса (apps.finances.repository.fifo_inputs +
apps.finances.fifo.compute_fifo, тот же батч-паттерн, что
apps/dashboard/services.py::get_dashboard использует для deferred_total —
один проход по всем ученикам, без запроса в цикле).

«Посещено уроков за месяц» тоже выводится из inp['cons_by_key']
(fifo_inputs()), а не отдельным запросом к LessonAttendance: так «посещено»
и «отработано» относятся к одному и тому же месяцу для одного и того же
события. И доп.урок, и сгорание — отдельные записи-уроки (lesson_type
'extra'/'burned'), их собственная lesson_date = дата проведения/сжигания
естественно относит деньги к нужному месяцу, без отдельного ремаппинга.

См. docs/superpowers/specs/2026-07-15-accounting-monthly-report-design.md
"""
from __future__ import annotations

import datetime
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from apps.core.utils.dates import msk_month_range
from apps.core.utils.decimal import js_number
from apps.finances.fifo import compute_fifo
from apps.finances.repository import (
    _date_str,
    balances_for_students,
    fifo_inputs,
    free_attended_units_by_month,
)
from apps.payments.models import Payment
from apps.students.models import Student


@dataclass
class MonthlyReportRow:
    student_id: int
    full_name: str
    platform_id: str | None
    attended_lessons: int | float          # оплаченные + бесплатные посещённые
    free_attended_lessons: int | float     # из них бесплатных (present, is_free)
    worked_off_month: Decimal
    unit_prices_month: list[Decimal]
    unit_qtys_month: list[Decimal]          # уроков по каждой цене (выровнено с prices)
    payments: list[tuple[str, Decimal]]
    paid_month_total: Decimal
    balance: int | float
    remaining_value: Decimal


def collect_monthly_report(month: str) -> list[MonthlyReportRow]:
    """
    Данные отчёта по ВСЕМ ученикам системы за указанный месяц.

    month: 'YYYY-MM'. Посещаемость/оплаты (kind purchase+extra)/отработанные деньги —
    только внутри месяца [month_start, month_end] включительно. Баланс и остаток
    аванса — на сегодня (те же величины, что apps.finances.balance.get_student_balance
    отдаёт per-student, но здесь батчево на всех разом).

    Raises:
        ValueError: month не в формате YYYY-MM / невалидный месяц (1-12).
    """
    month_start, month_end = msk_month_range(f'{month}-01')
    # compute_fifo работает с эксклюзивной верхней границей [start, end) —
    # msk_month_range отдаёт ВКЛЮЧИТЕЛЬНЫЙ последний день, поэтому +1 день.
    month_end_exclusive = (
        datetime.date.fromisoformat(month_end) + datetime.timedelta(days=1)
    ).strftime('%Y-%m-%d')

    students = list(
        Student.objects.order_by('full_name').values('id', 'full_name', 'platform_id')
    )
    student_ids = [s['id'] for s in students]

    # kind__in=['purchase','extra']: доплата за доп.урок сверх курса (kind='extra')
    # — реальная выручка и полноценная FIFO-партия (fifo_inputs берёт всё, кроме
    # refund), поэтому в листе оплат/«Итого оплачено» она обязана присутствовать —
    # иначе отчёт теряет деньги и рассинхронен с «Остатком аванса». refund не
    # входит: это возврат, он гасит остаток отдельной синтетической записью в FIFO.
    payment_rows = (
        Payment.objects
        .filter(kind__in=['purchase', 'extra'], paid_at__gte=month_start, paid_at__lte=month_end)
        .order_by('student_id', 'paid_at', 'id')
        .values('student_id', 'paid_at', 'total_amount')
    )
    payments_by_student: dict[int, list[tuple[str, Decimal]]] = {}
    for r in payment_rows:
        payments_by_student.setdefault(r['student_id'], []).append(
            (_date_str(r['paid_at']), r['total_amount'])
        )

    balances = balances_for_students(student_ids)

    # Бесплатные посещённые уроки за месяц — отдельным батчем: в FIFO-потребление
    # они не входят (is_free=False там), но «Посещено» их должно учитывать.
    free_by_student = free_attended_units_by_month(month_start, month_end)

    inp = fifo_inputs()
    remaining_value_by_student: dict[int, Decimal] = {}
    worked_off_month_by_student: dict[int, Decimal] = {}
    unit_prices_month_by_student: dict[int, list[Decimal]] = {}
    unit_qtys_month_by_student: dict[int, list[Decimal]] = {}
    attended_by_student: dict[int, Decimal] = {}
    for key in inp['keys']:
        fifo = compute_fifo(
            inp['lots_by_key'].get(key, []), inp['cons_by_key'].get(key, []),
            month_start, month_end_exclusive,
        )
        sid = int(key)
        remaining_value_by_student[sid] = fifo['remaining_value']
        worked_off_month_by_student[sid] = fifo['worked_off_month']
        unit_prices_month_by_student[sid] = fifo['worked_off_unit_prices_month']
        unit_qtys_month_by_student[sid] = fifo['worked_off_units_month']

        # «Посещено уроков за месяц» — из ТЕХ ЖЕ consumption-записей, что и
        # «отработано» (inp['cons_by_key'], построены fifo_inputs()). В новой
        # модели факт доп.урока сам является consumption-записью в дату своего
        # проведения (исходный пропуск остаётся present=false), поэтому «посещено»
        # и «отработано» в отчёте всегда относятся к одному и тому же месяцу для
        # одного и того же события, без задвоения по доп.уроку.
        attended = Decimal('0')
        for c in inp['cons_by_key'].get(key, []):
            if c.get('refund'):
                continue
            # Инклюзивная граница [month_start, month_end] — семантически то же
            # окно, что compute_fifo(..., month_start, month_end_exclusive) выше
            # (month_end_exclusive = month_end + 1 день), просто без открытого
            # интервала: обе стороны должны двигаться синхронно при правке.
            if month_start <= c['date'] <= month_end:
                attended += c['units']
        attended_by_student[sid] = attended

    rows: list[MonthlyReportRow] = []
    for s in students:
        sid = s['id']
        payments = payments_by_student.get(sid, [])
        # «Посещено» = оплаченные (attended_by_student, из FIFO-потребления, без free)
        # + бесплатные (free_by_student). free показывается отдельной колонкой, поэтому
        # оплаченные = attended_lessons − free_attended_lessons сходятся с суммой
        # unit_qtys_month (детализация по ценам считает только оплаченные списания).
        paid_attended = attended_by_student.get(sid, Decimal('0'))
        free_attended = free_by_student.get(sid, Decimal('0'))
        rows.append(MonthlyReportRow(
            student_id=sid,
            full_name=s['full_name'],
            platform_id=s['platform_id'],
            attended_lessons=js_number(paid_attended + free_attended),
            free_attended_lessons=js_number(free_attended),
            worked_off_month=worked_off_month_by_student.get(sid, Decimal('0')),
            unit_prices_month=unit_prices_month_by_student.get(sid, []),
            unit_qtys_month=unit_qtys_month_by_student.get(sid, []),
            payments=payments,
            paid_month_total=sum((amount for _, amount in payments), Decimal('0')),
            balance=balances.get(sid, 0),
            remaining_value=remaining_value_by_student.get(sid, Decimal('0')),
        ))
    return rows


def _pair_cols(base: int, i: int) -> tuple[int, int]:
    """1-й/2-й столбец i-й пары (0-based) начиная с колонки base."""
    return base + 2 * i, base + 2 * i + 1


def build_report_workbook(rows: list[MonthlyReportRow]):
    """Собрать openpyxl.Workbook отчёта (один ученик = одна строка), без сохранения.

    Общее ядро для write_report_xlsx (запись в файл, CLI-команда) и
    render_report_bytes (байты для ReportJob в разделе «Отчёты»)."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    max_prices = max((len(r.unit_prices_month) for r in rows), default=0)
    max_payments = max((len(r.payments) for r in rows), default=0)

    # Фиксированные колонки: ФИО(1) / Platform ID(2) / Посещено(3) /
    # в т.ч. бесплатных(4) / Отработано ₽(5). Дальше — блок пар «Стоимость 1
    # урока i / Уроков по цене i» (2 колонки на цену), затем блок оплат (пары
    # дата/платёж), затем итог/баланс/остаток.
    free_col = 4
    worked_col = 5
    price_base = 6
    payments_base = price_base + max_prices * 2
    total_col = payments_base + max_payments * 2
    balance_col = total_col + 1
    remaining_col = balance_col + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    headers = [
        'ФИО ученика', 'Platform ID', 'Посещено уроков за месяц',
        'в т.ч. бесплатных', 'Отработано деньгами за месяц, ₽',
    ]
    for i in range(1, max_prices + 1):
        headers += [f'Стоимость 1 урока {i}', f'Уроков по цене {i}']
    for i in range(1, max_payments + 1):
        headers += [f'Дата {i}', f'Платёж {i}']
    headers += ['Итого оплачено за месяц, ₽', 'Остаток оплаченных уроков', 'Остаток аванса, ₽']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    money_fmt = '#,##0.00'
    date_fmt = 'DD.MM.YYYY'

    for row in rows:
        values: list = [
            row.full_name, row.platform_id or '-', row.attended_lessons,
            row.free_attended_lessons, float(row.worked_off_month),
        ]
        for i in range(max_prices):
            if i < len(row.unit_prices_month):
                values += [float(row.unit_prices_month[i]), float(row.unit_qtys_month[i])]
            else:
                values += ['-', '-']
        for i in range(max_payments):
            if i < len(row.payments):
                pay_date, pay_amount = row.payments[i]
                values += [datetime.date.fromisoformat(pay_date), float(pay_amount)]
            else:
                values += ['-', '-']
        values += [float(row.paid_month_total), row.balance, float(row.remaining_value)]
        ws.append(values)

    for excel_row in range(2, len(rows) + 2):
        ws.cell(row=excel_row, column=worked_col).number_format = money_fmt
        for i in range(max_prices):
            price_col, _qty_col = _pair_cols(price_base, i)
            ws.cell(row=excel_row, column=price_col).number_format = money_fmt
        for i in range(max_payments):
            date_col, amount_col = _pair_cols(payments_base, i)
            date_cell = ws.cell(row=excel_row, column=date_col)
            if isinstance(date_cell.value, datetime.date):
                date_cell.number_format = date_fmt
            ws.cell(row=excel_row, column=amount_col).number_format = money_fmt
        ws.cell(row=excel_row, column=total_col).number_format = money_fmt
        ws.cell(row=excel_row, column=remaining_col).number_format = money_fmt

    widths = {
        1: 32, 2: 14, 3: 12, free_col: 14, worked_col: 16,
        total_col: 18, balance_col: 14, remaining_col: 14,
    }
    for i in range(max_prices):
        price_col, qty_col = _pair_cols(price_base, i)
        widths[price_col] = 14
        widths[qty_col] = 12
    for i in range(max_payments):
        date_col, amount_col = _pair_cols(payments_base, i)
        widths[date_col] = 12
        widths[amount_col] = 12
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    return wb


def write_report_xlsx(rows: list[MonthlyReportRow], path: str | Path) -> None:
    """Пишет rows в один лист «Отчёт»: один ученик = одна строка (файл на диск)."""
    wb = build_report_workbook(rows)
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def render_report_bytes(rows: list[MonthlyReportRow]) -> bytes:
    """Отчёт как xlsx-байты (для хранения в ReportJob.content)."""
    import io
    wb = build_report_workbook(rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
