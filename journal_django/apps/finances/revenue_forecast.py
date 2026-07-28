"""
Прогноз признания выручки по месяцам («когда отработаем то, что уже внесли»).

Запрос бухгалтерии 2026-07-27: по каждому ученику разложить внесённые, но ещё
не отработанные деньги на календарные месяцы вперёд — сколько выручки какой
месяц признает.

Учётная природа: признание выручки по мере оказания услуг (ФСБУ 9/2020
«Доходы»). Пока уроки не проведены, оплата — аванс (кредиторская
задолженность), он же отложенная (незаработанная) выручка; по мере проведения
уроков аванс закрывается и становится выручкой. Единая оплата за N абонементов
делится на N месяцев, чтобы не завышать выручку месяца оплаты. Неотработанный
остаток здесь — это и есть отложенная выручка, а раскладка по месяцам — график
её признания.

Модель (решения пользователя 2026-07-27):
  • берём НЕОТРАБОТАННЫЙ остаток ученика — FIFO-хвост партий-оплат
    (apps.finances.fifo.compute_fifo → remaining_lots), а не все оплаты подряд:
    уже отработанные деньги прогнозировать нечего;
  • раскладка идёт от МЕСЯЦА ФОРМИРОВАНИЯ отчёта, а не от дат оплат. Тем самым
    «запланированные, но не отработанные» деньги прошлых месяцев автоматически
    съезжают в хвост — отдельного переноса не нужно, это то же действие;
  • один календарный месяц = один абонемент = 4 урока. Хвост режется на месяцы
    по 4 урока; последний месяц бывает неполным;
  • разрез — ученик × направление ОПЛАТЫ, лист на направление. Курсы идут
    параллельно, поэтому и 4 урока в месяц считаются по каждому направлению
    отдельно (ученик с Minecraft и Blender отрабатывает 4+4).

Деньги месяца = Σ(уроков из партии × цена урока этой партии). Месяц может
попасть на границу двух партий с разными ценами — поэтому и нужен несвёрнутый
remaining_lots. Округление до копеек — один раз на месяц, невязка отдаётся
последнему месяцу, чтобы сумма строки в точности равнялась остатку аванса
(бухгалтерская точность, см. память feedback_financial_accounting_precision).

Half-lesson: считаем в УРОКАХ, а не в занятиях. У 45-минутной группы 4 урока
в месяц — это 8 занятий; на деньги это не влияет, цена привязана к уроку.

См. docs/superpowers/specs/2026-07-27-revenue-forecast-design.md
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from apps.core.utils.dates import msk_month_range, msk_today
from apps.core.utils.decimal import round_kopecks
from apps.directions.models import Direction
from apps.finances.fifo import compute_fifo
from apps.finances.repository import fifo_inputs
from apps.students.models import Student

# Один месяц = один абонемент = 4 урока. Инвариант данных: во всех 3829 оплатах
# lessons_count / subscriptions_count == 4.
LESSONS_PER_MONTH = Decimal('4')

NO_DIRECTION = 'Без направления'

_MONTH_NAMES = (
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
)


def month_label(ym: str) -> str:
    """'2026-07' → 'Июль 2026'."""
    year, month = int(ym[:4]), int(ym[5:7])
    return f'{_MONTH_NAMES[month - 1]} {year}'


def next_month(ym: str) -> str:
    """'2026-12' → '2027-01'."""
    year, month = int(ym[:4]), int(ym[5:7])
    return f'{year + 1}-01' if month == 12 else f'{year}-{month + 1:02d}'


@dataclass
class ForecastRow:
    """Ученик × направление: остаток и его раскладка по месяцам."""

    student_id: int
    full_name: str
    direction_id: int | None
    direction_name: str
    remaining_lessons: Decimal
    # Остаток аванса = сумма ПЛАНОВЫХ месяцев (отложенная выручка).
    remaining_value: Decimal
    # Уже признанная выручка = сумма ФАКТИЧЕСКИХ месяцев. Всегда 0 в режиме
    # без истории.
    worked_off_value: Decimal = Decimal('0')
    # Цены урока партий, из которых состоит остаток (в порядке очереди).
    unit_prices: list[Decimal] = field(default_factory=list)
    # 'YYYY-MM' → сумма месяца (факт + план). Σ == worked_off_value + remaining_value.
    by_month: dict[str, Decimal] = field(default_factory=dict)
    # Месяцы, в которых есть ФАКТ отработки. Стартовый месяц может быть и в
    # fact_months, и содержать плановый добор — тогда ячейка смешанная.
    fact_months: set[str] = field(default_factory=set)


@dataclass
class Forecast:
    """Готовые данные отчёта: месяцы-колонки и строки."""

    start_month: str
    months: list[str]
    rows: list[ForecastRow]
    full_history: bool = False

    def month_kind(self, ym: str) -> str:
        """'факт' / 'план' / 'факт+план' — для подписи колонки."""
        if not self.full_history or ym > self.start_month:
            return 'план'
        if ym < self.start_month:
            return 'факт'
        return 'факт+план'


def _split_into_months(
    lots: list[dict],
    start_month: str,
    first_month_capacity: Decimal | None = None,
) -> tuple[dict[str, Decimal], Decimal]:
    """
    Разрезать хвост партий на месяцы по LESSONS_PER_MONTH уроков.

    first_month_capacity — сколько уроков ещё влезает в СТАРТОВЫЙ месяц. Нужен,
    потому что часть стартового месяца обычно уже отработана: если ученик в июле
    сходил на 2 урока, из аванса в июле остаётся добрать только 2, иначе июль
    посчитается как 4 факт-урока + 4 план-урока. None = месяц пустой, влезает
    полный абонемент. Ноль и меньше → стартовый месяц пропускается целиком.

    Возвращает ({'YYYY-MM': сумма}, всего уроков). Суммы округлены до копеек,
    невязка округления отдана последнему месяцу: Σ месяцев == округлённый
    остаток строки, иначе бухгалтерия не сойдётся по горизонтали.
    """
    by_month: dict[str, Decimal] = {}
    exact_total = Decimal('0')
    total_lessons = Decimal('0')

    idx = 0
    lot_left = lots[0]['lessons'] if lots else Decimal('0')
    ym = start_month
    capacity = LESSONS_PER_MONTH if first_month_capacity is None else first_month_capacity
    if capacity <= 0 and lots:
        # Стартовый месяц уже выбран фактом — прогноз начинается со следующего.
        ym = next_month(ym)
        capacity = LESSONS_PER_MONTH
    while idx < len(lots):
        month_capacity = capacity     # у стартового месяца бывает неполной
        capacity = LESSONS_PER_MONTH
        need = month_capacity
        month_exact = Decimal('0')
        while need > 0 and idx < len(lots):
            if lot_left <= 0:
                idx += 1
                if idx >= len(lots):
                    break
                lot_left = lots[idx]['lessons']
                continue
            take = need if need < lot_left else lot_left   # min(need, lot_left)
            month_exact += take * lots[idx]['price_per_lesson']
            total_lessons += take
            lot_left -= take
            need -= take
        if need < month_capacity:     # в этот месяц реально что-то легло
            by_month[ym] = round_kopecks(month_exact)
            exact_total += month_exact
            ym = next_month(ym)

    if by_month:
        # Невязка «сумма округлённых месяцев» vs «округлённый точный итог».
        last = list(by_month)[-1]
        residue = round_kopecks(exact_total) - sum(by_month.values(), Decimal('0'))
        if residue:
            by_month[last] += residue
    return by_month, total_lessons


def collect_forecast(month: str | None = None, full_history: bool = False) -> Forecast:
    """
    Раскладка денег ученика по месяцам.

    month: 'YYYY-MM'; None — текущий МСК-месяц. Это месяц, С КОТОРОГО идёт
    прогноз («на момент формирования отчёта»).

    full_history=False — только будущее: строки с ненулевым остатком аванса,
    колонки от start_month вперёд.
    full_history=True — вся история денег: прошлые месяцы фактом отработки,
    будущие прогнозом. Попадают и ученики, у которых остаток уже нулевой, но
    история есть.

    В обоих режимах стартовый месяц добирается до 4 уроков с учётом уже
    отработанного в нём: иначе месяц, в котором ученик уже сходил на занятия,
    посчитался бы и фактом, и полным плановым абонементом.

    Raises:
        ValueError: month не в формате YYYY-MM / невалидный месяц (1-12).
    """
    if month is None:
        month = msk_today()[:7]             # «месяц формирования отчёта», МСК
    else:
        msk_month_range(f'{month}-01')      # валидация формата, как в бухотчёте
    start_month = month

    names = dict(Student.objects.values_list('id', 'full_name'))
    direction_names = dict(Direction.objects.values_list('id', 'name'))

    inp = fifo_inputs()
    rows: list[ForecastRow] = []
    for key in inp['keys']:
        student_id = int(key)
        fifo = compute_fifo(
            inp['lots_by_key'].get(key, []), inp['cons_by_key'].get(key, []),
            '0001-01-01', '9999-12-31',      # окно «месяца» здесь не используется
        )
        # Факт отработки по (месяц, направление ПАРТИИ) — тот же разрез, что у
        # хвоста, иначе строка «ученик × направление» не сойдётся.
        fact = fifo['worked_off_by_month_lot_direction']

        # Хвост режем ПО НАПРАВЛЕНИЯМ: курсы идут параллельно, у каждого свои
        # 4 урока в месяц. Порядок партий внутри направления сохраняется — от
        # него зависит, по какой цене уйдут ближайшие месяцы.
        tail_by_direction: dict[int | None, list[dict]] = {}
        for lot in fifo['remaining_lots']:
            tail_by_direction.setdefault(lot['direction_id'], []).append(lot)

        directions = set(tail_by_direction)
        if full_history:
            directions |= {d for (_ym, d) in fact}

        for direction_id in directions:
            lots = tail_by_direction.get(direction_id, [])
            # Сколько уроков стартового месяца уже закрыто фактом — на столько
            # меньше прогноз кладёт в этот месяц.
            done_now = fact.get((start_month, direction_id), {}).get('lessons', Decimal('0'))
            plan, lessons = _split_into_months(
                lots, start_month, first_month_capacity=LESSONS_PER_MONTH - done_now,
            )
            history = (
                {ym: v['value'] for (ym, d), v in fact.items()
                 if d == direction_id and ym < start_month}
                if full_history else {}
            )
            if full_history:
                # Факт стартового месяца тоже история — он складывается с добором
                # плана в ту же ячейку.
                current_fact = fact.get((start_month, direction_id), {}).get('value')
                if current_fact:
                    history[start_month] = current_fact
            if not plan and not history:
                continue

            by_month = dict(sorted(
                {ym: history.get(ym, Decimal('0')) + plan.get(ym, Decimal('0'))
                 for ym in set(history) | set(plan)}.items()
            ))
            rows.append(ForecastRow(
                student_id=student_id,
                full_name=names.get(student_id, f'#{student_id}'),
                direction_id=direction_id,
                direction_name=(direction_names.get(direction_id, f'#{direction_id}')
                                if direction_id is not None else NO_DIRECTION),
                remaining_lessons=lessons,
                remaining_value=sum(plan.values(), Decimal('0')),
                worked_off_value=sum(history.values(), Decimal('0')),
                unit_prices=list(dict.fromkeys(lot['price_per_lesson'] for lot in lots)),
                by_month=by_month,
                fact_months=set(history),
            ))

    rows.sort(key=lambda r: (r.direction_name, r.full_name))

    # Колонки: только месяцы, где реально есть деньги. Сплошной ряд не годится —
    # в данных есть выброс (одна запись 2023-01 при истории с 2025-01), он дал бы
    # два десятка пустых столбцов.
    months = sorted({m for r in rows for m in r.by_month})
    return Forecast(start_month=start_month, months=months, rows=rows,
                    full_history=full_history)


# ---------------------------------------------------------------------------
# Рендер
# ---------------------------------------------------------------------------

_INVALID_SHEET_CHARS = set(r'[]:*?/\\')


def _sheet_title(name: str, used: set[str]) -> str:
    """Имя листа Excel: ≤31 символа, без запрещённых символов, уникальное."""
    clean = ''.join(' ' if ch in _INVALID_SHEET_CHARS else ch for ch in name).strip()
    clean = (clean or 'Лист')[:31]
    title, n = clean, 2
    while title in used:
        suffix = f' ({n})'
        title = clean[:31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def build_workbook(forecast: Forecast):
    """openpyxl.Workbook: лист «Сводка» + лист на каждое направление."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    money_fmt = '#,##0.00'
    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='EDEDED')
    # Факт визуально отделён от плана: бухгалтерии важно не перепутать уже
    # признанную выручку с прогнозной.
    fact_fill = PatternFill('solid', fgColor='DCE9D5')
    plan_fill = PatternFill('solid', fgColor='FDE9D9')
    total_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')

    wb = openpyxl.Workbook()
    used_titles: set[str] = set()

    def month_headers() -> list[str]:
        if not forecast.full_history:
            return [month_label(m) for m in forecast.months]
        return [f'{month_label(m)} · {forecast.month_kind(m)}' for m in forecast.months]

    def style_header(ws, first_month_col: int) -> None:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        if not forecast.full_history:
            return
        for i, ym in enumerate(forecast.months):
            kind = forecast.month_kind(ym)
            ws.cell(row=1, column=first_month_col + i).fill = (
                fact_fill if kind == 'факт' else plan_fill
            )

    # В режиме полной истории добавляется колонка уже признанной выручки; из-за
    # неё месяцы начинаются на столбец правее.
    recognised_col = 5 if forecast.full_history else None
    first_month_col = 6 if forecast.full_history else 5

    def _sums(rows_: list[ForecastRow]) -> list:
        """Хвост строки: [признано] + суммы по месяцам."""
        out: list = []
        if forecast.full_history:
            out.append(float(sum((r.worked_off_value for r in rows_), Decimal('0'))))
        out += [float(sum((r.by_month.get(m, Decimal('0')) for r in rows_), Decimal('0')))
                for m in forecast.months]
        return out

    def _finish(ws, money_from: int) -> None:
        """Формат денег, ширины, закрепление шапки."""
        for cell in ws[ws.max_row]:
            cell.font = total_font
        for r_idx in range(2, ws.max_row + 1):
            for c_idx in range(money_from, first_month_col + len(forecast.months)):
                if c_idx == 4 and money_from < 4:      # «Цена урока» — текст
                    continue
                ws.cell(row=r_idx, column=c_idx).number_format = money_fmt
        for c_idx in range(2, first_month_col + len(forecast.months)):
            ws.column_dimensions[get_column_letter(c_idx)].width = 17
        ws.freeze_panes = 'B2'

    # --- Сводка: направление × месяц (главное число для бухгалтерии) ----------
    summary = wb.active
    summary.title = _sheet_title('Сводка', used_titles)
    # «Остаток аванса, ₽» — та же формулировка, что в бухотчёте за месяц
    # (apps/finances/reports.py): это отложенная выручка, ещё не признанная.
    summary_head = ['Направление', 'Учеников', 'Остаток уроков', 'Остаток аванса, ₽']
    if recognised_col:
        summary_head.append('Признано выручки, ₽')
    summary.append(summary_head + month_headers())
    style_header(summary, first_month_col)

    directions: dict[str, list[ForecastRow]] = {}
    for row in forecast.rows:
        directions.setdefault(row.direction_name, []).append(row)

    for direction_name, drows in directions.items():
        summary.append([
            direction_name,
            len({r.student_id for r in drows}),
            float(sum((r.remaining_lessons for r in drows), Decimal('0'))),
            float(sum((r.remaining_value for r in drows), Decimal('0'))),
        ] + _sums(drows))

    summary.append(['ИТОГО', len({r.student_id for r in forecast.rows}),
                    float(sum((r.remaining_lessons for r in forecast.rows), Decimal('0'))),
                    float(sum((r.remaining_value for r in forecast.rows), Decimal('0')))]
                   + _sums(forecast.rows))
    summary.column_dimensions['A'].width = 28
    _finish(summary, money_from=4)

    # --- Лист на направление: строка = ученик --------------------------------
    for direction_name, drows in directions.items():
        ws = wb.create_sheet(_sheet_title(direction_name, used_titles))
        head = ['ФИО ученика', 'Остаток уроков', 'Остаток аванса, ₽', 'Цена урока, ₽']
        if recognised_col:
            head.append('Признано выручки, ₽')
        ws.append(head + month_headers())
        style_header(ws, first_month_col)

        for row in drows:
            values: list = [
                row.full_name,
                float(row.remaining_lessons),
                float(row.remaining_value),
                # Остаток может состоять из партий с разной ценой — тогда их
                # несколько; строкой, чтобы не терять ни одну.
                '; '.join(f'{p:.2f}' for p in row.unit_prices),
            ]
            values += _sums([row])
            ws.append(values)

        ws.append(['ИТОГО',
                   float(sum((r.remaining_lessons for r in drows), Decimal('0'))),
                   float(sum((r.remaining_value for r in drows), Decimal('0'))), '']
                  + _sums(drows))
        ws.column_dimensions['A'].width = 32
        _finish(ws, money_from=3)

    return wb


def write_xlsx(forecast: Forecast, path: str | Path) -> None:
    """Пишет прогноз в файл на диск (CLI-команда export_revenue_forecast)."""
    wb = build_workbook(forecast)
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def render_bytes(forecast: Forecast) -> bytes:
    """Прогноз как xlsx-байты (для раздела «Отчёты»)."""
    import io
    wb = build_workbook(forecast)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
