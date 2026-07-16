"""
Тесты management command export_accounting_report — сквозной прогон
(call_command → читаем результат через openpyxl).
"""
from __future__ import annotations

from decimal import Decimal

import openpyxl
import pytest
from django.core.management import CommandError, call_command
from django.db import connection

pytestmark = pytest.mark.django_db


def _add_payment(created, student_id, direction_id, subs, total, paid_at):
    with connection.cursor() as cur:
        cur.execute(
            "INSERT INTO payments (student_id, direction_id, subscriptions_count, lessons_count, "
            "unit_price, total_amount, paid_at, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,'test') RETURNING id",
            [student_id, direction_id, subs, subs * 4, total, total, paid_at],
        )
        pid = cur.fetchone()[0]
    created['payments'].append(pid)
    return pid


def test_command_writes_xlsx_with_student_row(
    student_fixture, direction_fixture, graph_cleanup, tmp_path,
):
    _add_payment(graph_cleanup, student_fixture, direction_fixture, 1, 2000, '2026-07-05')
    out_path = tmp_path / 'out.xlsx'

    call_command('export_accounting_report', month='2026-07', out=str(out_path))

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[:3] == ['ФИО ученика', 'Platform ID', 'Посещено уроков за месяц']
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert '__fin_student__' in names


def test_command_invalid_month_raises_command_error(tmp_path):
    with pytest.raises(CommandError):
        call_command('export_accounting_report', month='2026-13', out=str(tmp_path / 'x.xlsx'))
    assert not (tmp_path / 'x.xlsx').exists()


def test_command_default_out_path_uses_reports_dir(settings, tmp_path, monkeypatch):
    settings.BASE_DIR = tmp_path
    monkeypatch.chdir(tmp_path)

    call_command('export_accounting_report', month='2026-07')

    expected = tmp_path / 'reports' / 'accounting_report_2026-07.xlsx'
    assert expected.exists()
