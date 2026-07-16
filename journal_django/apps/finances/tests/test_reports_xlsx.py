"""
Тесты чистой функции записи Excel (apps/finances/reports.py::write_report_xlsx).
Без БД — строки собираются вручную.
"""
from __future__ import annotations

import datetime
from decimal import Decimal

import openpyxl

from apps.finances.reports import MonthlyReportRow, write_report_xlsx


def test_write_report_xlsx_dynamic_price_and_payment_columns_and_dashes(tmp_path):
    rows = [
        MonthlyReportRow(
            student_id=1, full_name='Аня А.', platform_id='PL-1',
            attended_lessons=2, worked_off_month=Decimal('1000.00'),
            unit_prices_month=[Decimal('500.00')],
            payments=[('2026-07-05', Decimal('2000.00'))],
            paid_month_total=Decimal('2000.00'), balance=6, remaining_value=Decimal('2000.00'),
        ),
        MonthlyReportRow(
            student_id=2, full_name='Боря Б.', platform_id=None,
            attended_lessons=0, worked_off_month=Decimal('1450.00'),
            unit_prices_month=[Decimal('500.00'), Decimal('450.00')],
            payments=[('2026-07-01', Decimal('1000.00')), ('2026-07-15', Decimal('1500.00'))],
            paid_month_total=Decimal('2500.00'), balance=8, remaining_value=Decimal('2500.00'),
        ),
    ]
    out = tmp_path / 'report.xlsx'

    write_report_xlsx(rows, out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.title == 'Отчёт'
    header = [c.value for c in ws[1]]
    # 2 ученика: максимум 2 цены-за-урок и максимум 2 платежа за месяц.
    assert header == [
        'ФИО ученика', 'Platform ID', 'Посещено уроков за месяц', 'Отработано деньгами за месяц, ₽',
        'Стоимость 1 урока 1', 'Стоимость 1 урока 2',
        'Дата 1', 'Платёж 1', 'Дата 2', 'Платёж 2',
        'Итого оплачено за месяц, ₽', 'Остаток оплаченных уроков', 'Остаток аванса, ₽',
    ]

    row1 = [c.value for c in ws[2]]  # Аня — 1 ценовой сегмент и 1 платёж из 2 возможных
    assert row1[0:4] == ['Аня А.', 'PL-1', 2, 1000.0]
    assert row1[4] == 500.0
    assert row1[5] == '-'  # нет второго ценового сегмента
    # openpyxl всегда читает дату обратно как datetime.datetime (даже если
    # записан был datetime.date) — задокументированное поведение библиотеки,
    # т.к. формат ячеек Excel не различает "дату" и "дату-время".
    assert row1[6] == datetime.datetime(2026, 7, 5)
    assert row1[7] == 2000.0
    assert row1[8] == '-'  # нет второго платежа
    assert row1[9] == '-'
    assert row1[10:13] == [2000.0, 6, 2000.0]

    row2 = [c.value for c in ws[3]]  # Боря — platform_id пуст → '-', 2 ценовых сегмента, 2 платежа
    assert row2[1] == '-'
    assert row2[4] == 500.0
    assert row2[5] == 450.0
    assert row2[6] == datetime.datetime(2026, 7, 1)
    assert row2[8] == datetime.datetime(2026, 7, 15)


def test_write_report_xlsx_students_with_no_payments_this_month(tmp_path):
    """Есть ученики, но ни у кого нет платежей за месяц → нет пар "Дата N"/"Платёж N"."""
    rows = [
        MonthlyReportRow(
            student_id=1, full_name='Аня А.', platform_id='PL-1',
            attended_lessons=3, worked_off_month=Decimal('500.00'),
            unit_prices_month=[Decimal('500.00')], payments=[],
            paid_month_total=Decimal('0'), balance=1, remaining_value=Decimal('0'),
        ),
    ]
    out = tmp_path / 'no_payments.xlsx'

    write_report_xlsx(rows, out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == [
        'ФИО ученика', 'Platform ID', 'Посещено уроков за месяц', 'Отработано деньгами за месяц, ₽',
        'Стоимость 1 урока 1',
        'Итого оплачено за месяц, ₽', 'Остаток оплаченных уроков', 'Остаток аванса, ₽',
    ]
    row1 = [c.value for c in ws[2]]
    assert row1 == ['Аня А.', 'PL-1', 3, 500.0, 500.0, 0.0, 1, 0.0]


def test_write_report_xlsx_no_worked_off_no_payments(tmp_path):
    """Ни отработки, ни платежей за месяц → нет ни ценовых, ни платёжных колонок."""
    rows = [
        MonthlyReportRow(
            student_id=1, full_name='Аня А.', platform_id='PL-1',
            attended_lessons=0, worked_off_month=Decimal('0'),
            unit_prices_month=[], payments=[],
            paid_month_total=Decimal('0'), balance=1, remaining_value=Decimal('0'),
        ),
    ]
    out = tmp_path / 'no_activity.xlsx'

    write_report_xlsx(rows, out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == [
        'ФИО ученика', 'Platform ID', 'Посещено уроков за месяц', 'Отработано деньгами за месяц, ₽',
        'Итого оплачено за месяц, ₽', 'Остаток оплаченных уроков', 'Остаток аванса, ₽',
    ]
    row1 = [c.value for c in ws[2]]
    assert row1 == ['Аня А.', 'PL-1', 0, 0.0, 0.0, 1, 0.0]


def test_write_report_xlsx_empty_rows_writes_header_only(tmp_path):
    out = tmp_path / 'empty.xlsx'

    write_report_xlsx([], out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1
    assert ws['A1'].value == 'ФИО ученика'


def test_write_report_xlsx_creates_parent_dir(tmp_path):
    out = tmp_path / 'nested' / 'dir' / 'report.xlsx'

    write_report_xlsx([], out)

    assert out.exists()
