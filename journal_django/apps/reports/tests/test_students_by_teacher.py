"""
Тесты отчёта «Ученики по преподавателям за месяц»
(apps.reports.builders.students_by_teacher).

Лежат в apps/reports/tests намеренно: этот пакет НЕ переопределяет
django_db_setup (см. conftest.py рядом) → pytest-django поднимает свежую
мигрированную test_journal_test, где состав учеников и групп детерминирован.
Отчёт читает все активные членства базы, поэтому на общей persistent
journal_test проверки вида «в отчёте ровно N строк» были бы недостоверны.

См. docs/superpowers/specs/2026-09-03-students-by-teacher-report-design.md
"""
from __future__ import annotations

import datetime
import io
from decimal import Decimal

import openpyxl
import pytest

from apps.directions.models import Direction
from apps.groups.models import Group
from apps.lessons.models import Lesson, LessonAttendance
from apps.memberships.models import GroupMembership
from apps.reports import services
from apps.reports.builders.students_by_teacher import (
    build,
    collect_month,
    render_bytes,
)
from apps.reports.models import ReportType
from apps.students.models import Student
from apps.teachers.models import Teacher

pytestmark = pytest.mark.django_db

CREATED_AT = datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC)
MONTH = '2026-07'


@pytest.fixture
def data():
    """Фабрика доменных объектов отчёта (ORM, свежая test-БД чистится сама)."""

    class F:
        def __init__(self):
            self.direction = Direction.objects.create(
                name='__sbt_dir__', total_lessons=8, active=True,
            )

        def teacher(self, name: str) -> Teacher:
            return Teacher.objects.create(name=name, created_at=CREATED_AT)

        def group(self, name: str, teacher: Teacher, duration: int = 90) -> Group:
            return Group.objects.create(
                name=name, direction=self.direction, teacher=teacher,
                is_individual=False, lesson_duration_minutes=duration,
                created_at=CREATED_AT,
            )

        def student(self, full_name: str) -> Student:
            return Student.objects.create(full_name=full_name, created_at=CREATED_AT)

        def membership(self, student: Student, group: Group, active: bool = True):
            return GroupMembership.objects.create(
                student=student, group=group, active=active,
            )

        def lesson(self, group: Group, date: str, number: float,
                   lesson_type: str = 'regular', duration: int | None = None) -> Lesson:
            # submitted_by_token разводит уроки по типу: у пропуска и его
            # отработки совпадают дата+группа+номер, а natural key требует различия.
            return Lesson.objects.create(
                group=group, teacher=group.teacher, lesson_date=date,
                lesson_number=number,
                lesson_duration_minutes=(
                    duration if duration is not None else group.lesson_duration_minutes
                ),
                lesson_type=lesson_type,
                submitted_at=CREATED_AT,
                submitted_by_token=f'__sbt_test_{lesson_type}__',
            )

        def attend(self, lesson: Lesson, student: Student, present: bool,
                   is_free: bool = False, unpaid_skip: bool = False):
            return LessonAttendance.objects.create(
                lesson=lesson, student=student, present=present,
                is_free=is_free, unpaid_skip=unpaid_skip,
            )

    return F()


def _rows(month: str = MONTH, prefix: str = '__sbt_'):
    """Строки отчёта, ограниченные объектами теста (в свежей БД могут быть
    данные миграций)."""
    return [r for r in collect_month(month) if r.student_name.startswith(prefix)]


def _row(rows, student_name: str, group_name: str | None = None):
    found = [r for r in rows
             if r.student_name == student_name
             and (group_name is None or r.group_name == group_name)]
    assert len(found) == 1, f'ожидалась 1 строка, найдено {len(found)}'
    return found[0]


def test_counts_attended_lessons_of_the_month(data):
    """Считаются посещения месяца; пропуск и занятие соседнего месяца — нет."""
    t = data.teacher('__sbt_t1__')
    g = data.group('__sbt_g1__', t)
    s = data.student('__sbt_s1__')
    data.membership(s, g)
    data.attend(data.lesson(g, '2026-07-07', 1), s, present=True)
    data.attend(data.lesson(g, '2026-07-14', 2), s, present=False)
    data.attend(data.lesson(g, '2026-07-21', 3), s, present=True)
    data.attend(data.lesson(g, '2026-08-04', 4), s, present=True)   # другой месяц
    data.attend(data.lesson(g, '2026-06-30', 0), s, present=True)   # другой месяц

    row = _row(_rows(), '__sbt_s1__')

    assert (row.group_name, row.teacher_name) == ('__sbt_g1__', '__sbt_t1__')
    assert row.lessons == Decimal('2')


def test_half_lesson_weight(data):
    """45-минутное занятие = 0.5 урока; вес берётся с занятия, не с группы."""
    t = data.teacher('__sbt_t2__')
    g = data.group('__sbt_g2__', t, duration=45)
    s = data.student('__sbt_s2__')
    data.membership(s, g)
    for i, day in enumerate(('2026-07-06', '2026-07-08', '2026-07-13'), start=1):
        data.attend(data.lesson(g, day, i), s, present=True)
    # Отработка своей длительности (90) — вес 1, хотя группа 45-минутная.
    data.attend(data.lesson(g, '2026-07-20', 3.5, lesson_type='extra', duration=90),
                s, present=True)

    assert _row(_rows(), '__sbt_s2__').lessons == Decimal('2.5')


def test_burned_lesson_is_not_attendance(data):
    """Сгорание материализовано как present=true — но посещением не считается."""
    t = data.teacher('__sbt_t3__')
    g = data.group('__sbt_g3__', t)
    s = data.student('__sbt_s3__')
    data.membership(s, g)
    data.attend(data.lesson(g, '2026-07-07', 1), s, present=True)
    missed = data.lesson(g, '2026-07-14', 2)
    data.attend(missed, s, present=False)
    data.attend(data.lesson(g, '2026-07-14', 2, lesson_type='burned'), s, present=True)

    assert _row(_rows(), '__sbt_s3__').lessons == Decimal('1')


def test_extra_and_free_lessons_count(data):
    """Отработка, доп.урок сверх курса и бесплатное занятие — реальное присутствие."""
    t = data.teacher('__sbt_t4__')
    g = data.group('__sbt_g4__', t)
    s = data.student('__sbt_s4__')
    data.membership(s, g)
    data.attend(data.lesson(g, '2026-07-07', 1, lesson_type='extra'), s, present=True)
    data.attend(data.lesson(g, '2026-07-14', 2), s, present=True, is_free=True)
    # Неоплаченный пропуск — present=false, в счёт не идёт.
    data.attend(data.lesson(g, '2026-07-21', 3), s, present=False, unpaid_skip=True)

    assert _row(_rows(), '__sbt_s4__').lessons == Decimal('2')


def test_student_in_two_groups_gives_two_rows(data):
    """Две группы одновременно — две строки, каждая со своим преподавателем."""
    t1, t2 = data.teacher('__sbt_t5a__'), data.teacher('__sbt_t5b__')
    g1, g2 = data.group('__sbt_g5a__', t1), data.group('__sbt_g5b__', t2)
    s = data.student('__sbt_s5__')
    data.membership(s, g1)
    data.membership(s, g2)
    data.attend(data.lesson(g1, '2026-07-07', 1), s, present=True)
    data.attend(data.lesson(g2, '2026-07-08', 1), s, present=True)
    data.attend(data.lesson(g2, '2026-07-15', 2), s, present=True)

    rows = [r for r in _rows() if r.student_name == '__sbt_s5__']

    assert [(r.group_name, r.teacher_name, r.lessons) for r in rows] == [
        ('__sbt_g5a__', '__sbt_t5a__', Decimal('1')),
        ('__sbt_g5b__', '__sbt_t5b__', Decimal('2')),
    ]


def test_active_membership_without_lessons_gives_zero_row(data):
    """Активное членство без единого посещения — строка с 0, не пропуск."""
    t = data.teacher('__sbt_t6__')
    g = data.group('__sbt_g6__', t)
    s = data.student('__sbt_s6__')
    data.membership(s, g)

    assert _row(_rows(), '__sbt_s6__').lessons == Decimal('0')


def test_closed_membership_with_lessons_stays_in_month(data):
    """Ученик ушёл после месяца — из отчёта того месяца он не исчезает."""
    t = data.teacher('__sbt_t7__')
    g = data.group('__sbt_g7__', t)
    s = data.student('__sbt_s7__')
    data.membership(s, g, active=False)
    data.attend(data.lesson(g, '2026-07-07', 1), s, present=True)

    assert _row(_rows(), '__sbt_s7__').lessons == Decimal('1')


def test_closed_membership_without_lessons_is_absent(data):
    """Закрытое членство без занятий месяца строки не даёт."""
    t = data.teacher('__sbt_t8__')
    g = data.group('__sbt_g8__', t)
    s = data.student('__sbt_s8__')
    data.membership(s, g, active=False)

    assert [r for r in _rows() if r.student_name == '__sbt_s8__'] == []


def test_student_without_group_is_absent(data):
    """Ученик без единой группы в отчёт не попадает — у него нет преподавателя."""
    data.student('__sbt_s9__')

    assert [r for r in _rows() if r.student_name == '__sbt_s9__'] == []


def test_group_month_lessons_counts_course_lessons_only(data):
    """«Уроков у группы» — курсовая сетка месяца: доп.урок и сгорание не в счёт."""
    t = data.teacher('__sbt_tE__')
    g = data.group('__sbt_gE__', t)
    s1, s2 = data.student('__sbt_sE1__'), data.student('__sbt_sE2__')
    data.membership(s1, g)
    data.membership(s2, g)
    for i, day in enumerate(('2026-07-07', '2026-07-14', '2026-07-21'), start=1):
        lesson = data.lesson(g, day, i)
        data.attend(lesson, s1, present=True)
        data.attend(lesson, s2, present=(i == 1))
    # Сетку группы не удлиняют: отработка адресная, сгорание — списание денег.
    data.attend(data.lesson(g, '2026-07-28', 3.5, lesson_type='extra'), s2, present=True)
    data.attend(data.lesson(g, '2026-07-14', 2, lesson_type='burned'), s2, present=True)
    # Занятие соседнего месяца в июльскую сетку тоже не входит.
    data.attend(data.lesson(g, '2026-08-04', 4), s1, present=True)

    rows = _rows()

    assert _row(rows, '__sbt_sE1__').group_lessons == Decimal('3')
    assert _row(rows, '__sbt_sE1__').lessons == Decimal('3')
    # У ученика 1 занятие группы + отработка; сгорание не считается.
    assert _row(rows, '__sbt_sE2__').group_lessons == Decimal('3')
    assert _row(rows, '__sbt_sE2__').lessons == Decimal('2')


def test_group_month_lessons_use_half_lesson_weight(data):
    """Сетка группы считается в уроках: 45-минутные занятия — по 0.5."""
    t = data.teacher('__sbt_tF__')
    g = data.group('__sbt_gF__', t, duration=45)
    s = data.student('__sbt_sF__')
    data.membership(s, g)
    for i, day in enumerate(('2026-07-06', '2026-07-08', '2026-07-13'), start=1):
        data.attend(data.lesson(g, day, i), s, present=(i != 2))

    row = _row(_rows(), '__sbt_sF__')

    assert (row.group_lessons, row.lessons) == (Decimal('1.5'), Decimal('1'))


def test_group_without_lessons_in_month_has_zero_group_lessons(data):
    """Группа не занималась в месяце — 0 уроков у группы, не пустая ячейка."""
    t = data.teacher('__sbt_tG__')
    g = data.group('__sbt_gG__', t)
    s = data.student('__sbt_sG__')
    data.membership(s, g)

    assert _row(_rows(), '__sbt_sG__').group_lessons == Decimal('0')


def test_direction_of_the_group_is_reported(data):
    """Направление — свойство группы ученика."""
    t = data.teacher('__sbt_tH__')
    g = data.group('__sbt_gH__', t)
    s = data.student('__sbt_sH__')
    data.membership(s, g)

    assert _row(_rows(), '__sbt_sH__').direction_name == '__sbt_dir__'


def test_rows_sorted_by_teacher_then_group_then_student(data):
    """Сортировка: преподаватель → группа → ученик."""
    t1, t2 = data.teacher('__sbt_tA__'), data.teacher('__sbt_tB__')
    g_a1 = data.group('__sbt_gA1__', t1)
    g_a2 = data.group('__sbt_gA2__', t1)
    g_b = data.group('__sbt_gB__', t2)
    s1, s2 = data.student('__sbt_sX__'), data.student('__sbt_sY__')
    for s, g in ((s1, g_a2), (s1, g_a1), (s2, g_a1), (s1, g_b)):
        data.membership(s, g)

    rows = [r for r in _rows() if r.teacher_name.startswith('__sbt_t')]

    assert [(r.teacher_name, r.group_name, r.student_name) for r in rows] == [
        ('__sbt_tA__', '__sbt_gA1__', '__sbt_sX__'),
        ('__sbt_tA__', '__sbt_gA1__', '__sbt_sY__'),
        ('__sbt_tA__', '__sbt_gA2__', '__sbt_sX__'),
        ('__sbt_tB__', '__sbt_gB__', '__sbt_sX__'),
    ]


def test_render_bytes_writes_sheet(data):
    """Лист: шапка, плоские строки, дробное число уроков числом."""
    t = data.teacher('__sbt_tC__')
    g = data.group('__sbt_gC__', t, duration=45)
    s = data.student('__sbt_sC__')
    data.membership(s, g)
    data.attend(data.lesson(g, '2026-07-07', 1), s, present=True)

    wb = openpyxl.load_workbook(io.BytesIO(render_bytes(_rows())))
    ws = wb.active

    assert ws.title == 'Ученики по преподавателям'
    assert [c.value for c in ws[1]] == [
        'Группа', 'Преподаватель', 'Ученик',
        'Уроков у группы за месяц', 'Посещено учеником', 'Направление',
    ]
    body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert ['__sbt_gC__', '__sbt_tC__', '__sbt_sC__', 0.5, 0.5, '__sbt_dir__'] in body


def test_build_returns_named_file(data):
    """build() отдаёт байты, число строк и имя файла с месяцем."""
    t = data.teacher('__sbt_tD__')
    g = data.group('__sbt_gD__', t)
    s = data.student('__sbt_sD__')
    data.membership(s, g)

    content, row_count, filename = build(MONTH)

    assert content[:2] == b'PK'  # xlsx — это zip
    assert row_count >= 1
    assert filename == 'students_by_teacher_2026-07.xlsx'


def test_service_dispatches_report_type(data):
    """Тип отчёта зарегистрирован в диспетчере build_report."""
    content, _row_count, filename = services.build_report(
        ReportType.STUDENTS_BY_TEACHER, {'month': MONTH})

    assert content[:2] == b'PK'
    assert filename == 'students_by_teacher_2026-07.xlsx'


def test_invalid_month_raises():
    """Кривой месяц — ValueError (services пометит задачу failure)."""
    with pytest.raises(ValueError):
        collect_month('2026-13')
