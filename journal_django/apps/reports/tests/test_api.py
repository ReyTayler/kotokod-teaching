"""API-тесты раздела «Отчёты»: RBAC + run→status→download через celery result
backend (в eager-режиме .delay() исполняется синхронно). Ничего не хранится в БД."""
from __future__ import annotations

import io

import pytest

pytestmark = pytest.mark.django_db

BASE = '/api/admin/reports'
RUN = f'{BASE}/renewals_month/run'


def test_run_returns_task_id(manager_client):
    resp = manager_client.post(RUN, {'year': 2026, 'month': 5}, format='json')
    assert resp.status_code == 202
    assert resp.json()['task_id']


def test_run_then_status_success(manager_client):
    task_id = manager_client.post(RUN, {'year': 2026, 'month': 5}, format='json').json()['task_id']
    st = manager_client.get(f'{BASE}/status/{task_id}')
    assert st.status_code == 200
    body = st.json()
    assert body['state'] == 'SUCCESS'
    assert body['filename'] == 'renewals_2026-05.xlsx'
    # байты в статусе НЕ отдаём
    assert 'content_b64' not in body


def test_run_then_download_xlsx(manager_client):
    task_id = manager_client.post(RUN, {'year': 2026, 'month': 5}, format='json').json()['task_id']
    resp = manager_client.get(f'{BASE}/download/{task_id}')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp['Content-Type']
    assert 'attachment' in resp['Content-Disposition']
    from openpyxl import load_workbook
    assert load_workbook(io.BytesIO(resp.getvalue())).active.cell(row=2, column=1).value == 'ФИО ученика'


def test_download_unknown_task_not_ready(manager_client):
    # неизвестный task_id → PENDING → «ещё не готов» (400)
    resp = manager_client.get(f'{BASE}/download/00000000-0000-0000-0000-000000000000')
    assert resp.status_code == 400


def test_run_rejects_bad_month(manager_client):
    assert manager_client.post(RUN, {'year': 2026, 'month': 13}, format='json').status_code == 400


def test_run_rejects_future_month(manager_client):
    assert manager_client.post(RUN, {'year': 2999, 'month': 12}, format='json').status_code == 400


def test_run_unknown_report_type_404(manager_client):
    assert manager_client.post(f'{BASE}/no_such/run', {'year': 2026, 'month': 5},
                               format='json').status_code == 404


def test_teacher_forbidden(teacher_client):
    assert teacher_client.post(RUN, {'year': 2026, 'month': 5}, format='json').status_code == 403
    assert teacher_client.get(f'{BASE}/status/x').status_code == 403


def test_anon_unauthorized(anon_client):
    assert anon_client.post(RUN, {'year': 2026, 'month': 5}, format='json').status_code in (401, 403)


def test_admin_allowed(admin_client):
    assert admin_client.post(RUN, {'year': 2026, 'month': 5}, format='json').status_code == 202


# ---------------------------------------------------------------------------
# Отчёт по посещаемости и прогноз отработки денег (добавлены 2026-07-28)
# ---------------------------------------------------------------------------

ATTENDANCE_RUN = f'{BASE}/attendance_month/run'
FORECAST_RUN = f'{BASE}/revenue_forecast/run'


def test_attendance_run_then_download(manager_client):
    task_id = manager_client.post(
        ATTENDANCE_RUN, {'month': '2026-05'}, format='json').json()['task_id']

    st = manager_client.get(f'{BASE}/status/{task_id}').json()
    assert st['state'] == 'SUCCESS'
    assert st['filename'] == 'attendance_2026-05.xlsx'

    resp = manager_client.get(f'{BASE}/download/{task_id}')
    assert resp.status_code == 200
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.getvalue())).active
    assert ws.cell(row=1, column=1).value == 'ФИО ученика'
    assert ws.cell(row=1, column=2).value == 'Группа'


def test_attendance_rejects_bad_and_future_month(manager_client):
    assert manager_client.post(ATTENDANCE_RUN, {'month': '2026-13'},
                               format='json').status_code == 400
    assert manager_client.post(ATTENDANCE_RUN, {'month': '2999-12'},
                               format='json').status_code == 400


def test_forecast_run_then_download(manager_client):
    task_id = manager_client.post(
        FORECAST_RUN, {'month': '2026-05'}, format='json').json()['task_id']

    st = manager_client.get(f'{BASE}/status/{task_id}').json()
    assert st['state'] == 'SUCCESS'
    assert st['filename'] == 'revenue_forecast_2026-05.xlsx'

    resp = manager_client.get(f'{BASE}/download/{task_id}')
    assert resp.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.getvalue()))
    assert wb.sheetnames[0] == 'Сводка'
    assert wb['Сводка'].cell(row=1, column=1).value == 'Направление'


def test_forecast_full_history_flag_changes_file_and_columns(manager_client):
    task_id = manager_client.post(
        FORECAST_RUN, {'month': '2026-05', 'full_history': True},
        format='json').json()['task_id']

    st = manager_client.get(f'{BASE}/status/{task_id}').json()
    assert st['filename'] == 'revenue_forecast_2026-05_full.xlsx'

    resp = manager_client.get(f'{BASE}/download/{task_id}')
    from openpyxl import load_workbook
    header = [c.value for c in load_workbook(io.BytesIO(resp.getvalue()))['Сводка'][1]]
    assert 'Признано выручки, ₽' in header


def test_forecast_full_history_defaults_to_false(manager_client):
    task_id = manager_client.post(
        FORECAST_RUN, {'month': '2026-05'}, format='json').json()['task_id']

    resp = manager_client.get(f'{BASE}/download/{task_id}')
    from openpyxl import load_workbook
    header = [c.value for c in load_workbook(io.BytesIO(resp.getvalue()))['Сводка'][1]]
    assert 'Признано выручки, ₽' not in header


def test_new_reports_respect_rbac(teacher_client, anon_client):
    assert teacher_client.post(ATTENDANCE_RUN, {'month': '2026-05'},
                               format='json').status_code == 403
    assert teacher_client.post(FORECAST_RUN, {'month': '2026-05'},
                               format='json').status_code == 403
    assert anon_client.post(FORECAST_RUN, {'month': '2026-05'},
                            format='json').status_code in (401, 403)


# ---------------------------------------------------------------------------
# «Отчёт по переходимости» — без параметров, по всей истории
# ---------------------------------------------------------------------------

RETENTION_RUN = f'{BASE}/retention/run'


def test_retention_runs_without_params(manager_client):
    """Период не передаётся: отчёт общий. Пустое тело обязано приниматься."""
    resp = manager_client.post(RETENTION_RUN, {}, format='json')

    assert resp.status_code == 202
    assert resp.json()['task_id']


def test_retention_downloads_xlsx(manager_client):
    from openpyxl import load_workbook

    task_id = manager_client.post(RETENTION_RUN, {}, format='json').json()['task_id']
    resp = manager_client.get(f'{BASE}/download/{task_id}')

    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(b''.join(resp.streaming_content)
                                 if resp.streaming else resp.content))
    assert 'Свод — преподаватели' in wb.sheetnames
    assert 'Свод — направления' in wb.sheetnames
