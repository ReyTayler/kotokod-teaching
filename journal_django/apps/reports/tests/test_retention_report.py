"""
Тесты «Отчёта по переходимости» (apps.reports.builders.retention).

Как и остальные тесты reports, идут по свежей мигрированной test_journal_test
(django_db_setup НЕ переопределён — см. conftest.py пакета).
"""
from __future__ import annotations

import datetime

import pytest
from django.db import connection

from apps.reports.builders import retention


@pytest.fixture
def world(renewals_fixture):
    """Мир отчёта: направления, преподаватели, группы, ученики, членства.

    Сделки создаём методами renewals_fixture, обвязку «кто у кого учится» —
    своими: именно она связывает сделку с преподавателем и направлением.
    """
    created = {'membership': [], 'group': [], 'teacher': [], 'direction': []}

    class W:
        def __init__(self, base):
            self.base = base
            self.pipeline = base.pipeline()
            self.won = base.stage(self.pipeline, 'renewed', 'Продлён', 'won', sort_order=9)
            self.lost = base.stage(self.pipeline, 'churned', 'Ушёл', 'lost', sort_order=10)
            self.open = base.stage(self.pipeline, 'awaiting_renewal', 'Ждём продление',
                                   'decision', sort_order=5)

        def teacher(self, name, is_service=False):
            with connection.cursor() as cur:
                cur.execute(
                    'INSERT INTO teachers (name, active, is_service, created_at) '
                    'VALUES (%s, true, %s, now()) RETURNING id', [name, is_service])
                tid = cur.fetchone()[0]
            created['teacher'].append(tid)
            return tid

        def direction(self, name):
            with connection.cursor() as cur:
                cur.execute(
                    'INSERT INTO directions (name, total_lessons, active) '
                    'VALUES (%s, 36, true) RETURNING id', [name])
                did = cur.fetchone()[0]
            created['direction'].append(did)
            return did

        def group(self, name, teacher_id, direction_id, active=True):
            with connection.cursor() as cur:
                cur.execute(
                    'INSERT INTO groups (name, direction_id, teacher_id, is_individual, '
                    'lesson_duration_minutes, active, lesson_number_offset) '
                    'VALUES (%s, %s, %s, false, 90, %s, 0) RETURNING id',
                    [name, direction_id, teacher_id, active])
                gid = cur.fetchone()[0]
            created['group'].append(gid)
            return gid

        def enrol(self, group_id, student_id, active=True):
            with connection.cursor() as cur:
                cur.execute(
                    'INSERT INTO group_memberships (group_id, student_id, lessons_done, active) '
                    'VALUES (%s, %s, 0, %s) RETURNING id', [group_id, student_id, active])
                created['membership'].append(cur.fetchone()[0])

        @staticmethod
        def days_ago(days: int) -> datetime.date:
            return datetime.date.today() - datetime.timedelta(days=days)

        def stuck_date(self):
            """Опорная дата зависшей сделки — заведомо за порогом."""
            return self.days_ago(retention.STUCK_AFTER_DAYS + 10)

        def fresh_date(self):
            """Опорная дата сделки «в работе» — заведомо внутри порога."""
            return self.days_ago(1)

    yield W(renewals_fixture)

    with connection.cursor() as cur:
        for mid in created['membership']:
            cur.execute('DELETE FROM group_memberships WHERE id = %s', [mid])
        for gid in created['group']:
            cur.execute('DELETE FROM groups WHERE id = %s', [gid])
        for tid in created['teacher']:
            cur.execute('DELETE FROM teachers WHERE id = %s', [tid])
        for did in created['direction']:
            cur.execute('DELETE FROM directions WHERE id = %s', [did])


@pytest.fixture
def enrolled(world, renewals_fixture):
    """Один преподаватель, одно направление, фабрика «ученик в группе»."""
    teacher = world.teacher('__ret_teacher__')
    direction = world.direction('__ret_dir__')
    group = world.group('__ret_group__', teacher, direction)

    def _student(name):
        sid = renewals_fixture.student(name)
        world.enrol(group, sid)
        return sid

    return _student


def _cycle(rows: list[dict], cycle: int) -> dict:
    return next(r for r in rows if r['cycle'] == cycle)


def _row(rows: list[dict], name: str) -> dict:
    return next(r for r in rows if r['name'] == name)


# ---------------------------------------------------------------------------
# Воронка по циклам
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_funnel_splits_cycle_outcomes(world, renewals_fixture, enrolled):
    """Четыре исхода на цикле: перешёл дальше, ушёл, завис, в работе."""
    a, b, c, d = (enrolled(f'__ret_f_{x}__') for x in 'abcd')
    renewals_fixture.deal(a, world.pipeline, world.won, 3, closed_at='2026-07-15')
    renewals_fixture.deal(b, world.pipeline, world.lost, 3, closed_at='2026-07-15')
    renewals_fixture.deal(c, world.pipeline, world.open, 3, due_at=world.stuck_date())
    renewals_fixture.deal(d, world.pipeline, world.open, 3, due_at=world.fresh_date())

    row = _cycle(retention.collect()['funnel'], 3)

    assert row['reached'] == 4
    assert row['advanced'] == 1
    assert row['lost'] == 1
    assert row['stuck'] == 1
    assert row['active'] == 1


@pytest.mark.django_db
def test_rate_counts_stuck_as_loss_and_excludes_active(world, renewals_fixture, enrolled):
    """
    Знаменатель — только РЕШЁННЫЕ сделки.

    Зависшая идёт в потери (ушедшего обычно просто перестают вести, отметки нет),
    а «в работе» не считается вовсе — иначе переход занижался бы тем сильнее,
    чем больше у преподавателя активных учеников.

    Здесь: 2 перешли, 1 завис, 3 в работе → 2/3, а НЕ 2/6 и не 2/2.
    """
    students = [enrolled(f'__ret_r_{i}__') for i in range(6)]
    for sid in students[:2]:
        renewals_fixture.deal(sid, world.pipeline, world.won, 5, closed_at='2026-07-15')
    renewals_fixture.deal(students[2], world.pipeline, world.open, 5, due_at=world.stuck_date())
    for sid in students[3:]:
        renewals_fixture.deal(sid, world.pipeline, world.open, 5, due_at=world.fresh_date())

    row = _cycle(retention.collect()['funnel'], 5)

    assert row['rate'] == pytest.approx(2 / 3)


@pytest.mark.django_db
def test_rate_is_none_when_nothing_decided(world, renewals_fixture, enrolled):
    """None, а не 0 %: «ещё никто не дошёл до решения» и «все ушли» —
    противоположные вещи, а на экране 0 % выглядит одинаково."""
    student = enrolled('__ret_none__')
    renewals_fixture.deal(student, world.pipeline, world.open, 7, due_at=world.fresh_date())

    assert _cycle(retention.collect()['funnel'], 7)['rate'] is None


@pytest.mark.django_db
def test_lessons_column_maps_cycle_to_course_position(world, renewals_fixture, enrolled):
    """Цикл × 4 урока: девятый цикл — это 36 уроков, конец стандартного курса.
    Без этой колонки провал на 9-м цикле не связать с завершением курса."""
    student = enrolled('__ret_lessons__')
    renewals_fixture.deal(student, world.pipeline, world.won, 9, closed_at='2026-07-15')

    assert _cycle(retention.collect()['funnel'], 9)['lessons_to'] == 36


# ---------------------------------------------------------------------------
# Разрезы по преподавателям и направлениям
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_bands_group_cycles_by_meaning(world, renewals_fixture, enrolled):
    """Цикл 3 → диапазон 1–4, цикл 9 → 5–9 (граница конца курса)."""
    student = enrolled('__ret_bands__')
    renewals_fixture.deal(student, world.pipeline, world.won, 3, closed_at='2026-07-15')
    renewals_fixture.deal(student, world.pipeline, world.lost, 9, closed_at='2026-07-15')

    bands = _row(retention.collect()['teachers'], '__ret_teacher__')['bands']

    assert bands['1–4 (1–16 уроков)']['advanced'] == 1
    assert bands['5–9 (17–36, до конца курса)']['lost'] == 1
    assert bands['1–4 (1–16 уроков)']['rate'] == 1.0
    assert bands['5–9 (17–36, до конца курса)']['rate'] == 0.0


@pytest.mark.django_db
def test_student_of_two_teachers_counts_for_both(world, renewals_fixture):
    """Осознанная неэксклюзивность: сделка привязана к ученику, а не к группе,
    поэтому разделить её между преподавателями данные не позволяют."""
    t1 = world.teacher('__ret_t1__')
    t2 = world.teacher('__ret_t2__')
    direction = world.direction('__ret_dir_two__')
    student = renewals_fixture.student('__ret_student_two__')
    world.enrol(world.group('__ret_g1__', t1, direction), student)
    world.enrol(world.group('__ret_g2__', t2, direction), student)
    renewals_fixture.deal(student, world.pipeline, world.won, 1, closed_at='2026-07-15')

    rows = retention.collect()['teachers']

    assert _row(rows, '__ret_t1__')['bands']['1–4 (1–16 уроков)']['advanced'] == 1
    assert _row(rows, '__ret_t2__')['bands']['1–4 (1–16 уроков)']['advanced'] == 1


@pytest.mark.django_db
def test_directions_grouped_through_groups(world, renewals_fixture):
    """Направление берётся через группы ученика — тем же путём, что преподаватель,
    иначе листы разошлись бы между собой."""
    teacher = world.teacher('__ret_t_dir__')
    d1 = world.direction('__ret_dir_a__')
    d2 = world.direction('__ret_dir_b__')
    student = renewals_fixture.student('__ret_student_dir__')
    world.enrol(world.group('__ret_g_a__', teacher, d1), student)
    world.enrol(world.group('__ret_g_b__', teacher, d2), student)
    renewals_fixture.deal(student, world.pipeline, world.won, 2, closed_at='2026-07-15')

    rows = retention.collect()['directions']
    band = '1–4 (1–16 уроков)'

    assert _row(rows, '__ret_dir_a__')['bands'][band]['advanced'] == 1
    assert _row(rows, '__ret_dir_b__')['bands'][band]['advanced'] == 1


@pytest.mark.django_db
def test_former_student_still_counted(world, renewals_fixture):
    """Неактивное членство и архивная группа считаются: ушедший ученик — часть
    истории переходимости, иначе показываем только выживших."""
    teacher = world.teacher('__ret_t_former__')
    direction = world.direction('__ret_dir_former__')
    group = world.group('__ret_g_former__', teacher, direction, active=False)
    student = renewals_fixture.student('__ret_student_former__')
    world.enrol(group, student, active=False)
    renewals_fixture.deal(student, world.pipeline, world.lost, 1, closed_at='2026-07-15')

    row = _row(retention.collect()['teachers'], '__ret_t_former__')

    assert row['students'] == 1
    assert row['bands']['1–4 (1–16 уроков)']['lost'] == 1


@pytest.mark.django_db
def test_service_teacher_sorted_last(world, renewals_fixture):
    """Служебная запись несёт ~80 % сделок школы — в общем зачёте она забивает
    живых, поэтому уходит вниз независимо от объёма."""
    service = world.teacher('__ret_service__', is_service=True)
    live = world.teacher('__ret_live__')
    direction = world.direction('__ret_dir_service__')
    # У служебной записи учеников БОЛЬШЕ — иначе тест прошёл бы и при обычной
    # сортировке по объёму, не различая флаг.
    for i in range(3):
        sid = renewals_fixture.student(f'__ret_s_service_{i}__')
        world.enrol(world.group(f'__ret_g_service_{i}__', service, direction), sid)
        renewals_fixture.deal(sid, world.pipeline, world.won, 1, closed_at='2026-07-15')
    sid = renewals_fixture.student('__ret_s_live__')
    world.enrol(world.group('__ret_g_live__', live, direction), sid)
    renewals_fixture.deal(sid, world.pipeline, world.won, 1, closed_at='2026-07-15')

    names = [r['name'] for r in retention.collect()['teachers']]

    assert names.index('__ret_service__') > names.index('__ret_live__')


# ---------------------------------------------------------------------------
# Зависшие сделки — рабочий список
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_stuck_list_puts_live_teachers_before_import_traces(world, renewals_fixture):
    """
    Ученики живых преподавателей сверху, следы импорта — вниз.

    У импортных сделок возраст 1300+ дней; сортируй мы только по нему, они заняли
    бы весь верх и список стал бы нерабочим, хотя разбираться там не с кем.
    Поэтому свежая живая сделка обязана стоять выше древней служебной.
    """
    service = world.teacher('__ret_stuck_service__', is_service=True)
    live = world.teacher('__ret_stuck_live__')
    direction = world.direction('__ret_dir_stuck__')

    old = renewals_fixture.student('__ret_stuck_old__')
    world.enrol(world.group('__ret_g_stuck_service__', service, direction), old)
    renewals_fixture.deal(old, world.pipeline, world.open, 4, due_at=world.days_ago(1300))

    recent = renewals_fixture.student('__ret_stuck_recent__')
    world.enrol(world.group('__ret_g_stuck_live__', live, direction), recent)
    renewals_fixture.deal(recent, world.pipeline, world.open, 4, due_at=world.stuck_date())

    rows = retention.collect()['stuck']
    names = [r['student_name'] for r in rows]

    assert names.index('__ret_stuck_recent__') < names.index('__ret_stuck_old__')
    assert _row_by_student(rows, '__ret_stuck_recent__')['service_only'] is False
    assert _row_by_student(rows, '__ret_stuck_old__')['service_only'] is True


def _row_by_student(rows: list[dict], name: str) -> dict:
    return next(r for r in rows if r['student_name'] == name)


@pytest.mark.django_db
def test_fresh_open_deal_is_not_stuck(world, renewals_fixture, enrolled):
    """Сделка внутри порога — «в работе», в рабочий список не попадает."""
    student = enrolled('__ret_fresh__')
    renewals_fixture.deal(student, world.pipeline, world.open, 2, due_at=world.fresh_date())

    assert all(r['student_name'] != '__ret_fresh__' for r in retention.collect()['stuck'])


# ---------------------------------------------------------------------------
# Книга целиком
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_build_produces_workbook_with_charts(world, renewals_fixture, enrolled):
    import io

    from openpyxl import load_workbook

    # Два цикла — иначе график строить не по чему (одна точка).
    student = enrolled('__ret_wb__')
    renewals_fixture.deal(student, world.pipeline, world.won, 1, closed_at='2026-06-15')
    renewals_fixture.deal(student, world.pipeline, world.won, 2, closed_at='2026-07-15')

    content, rows, filename = retention.build()

    assert filename.startswith('retention_') and filename.endswith('.xlsx')
    assert rows >= 2
    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames == [
        'Воронка по циклам', 'Циклы × преподаватели', 'Циклы × направления',
        'Зависшие сделки', 'Детализация',
    ]
    # Оба графика на главном листе: кривая дожития и доля перехода — отдельными
    # диаграммами, а не двумя осями одной (см. докстринг модуля).
    assert len(wb['Воронка по циклам']._charts) == 2


@pytest.mark.django_db
def test_rate_column_is_a_formula(world, renewals_fixture, enrolled):
    """Переход считается формулой, а не записанным числом: лист обязан
    пересчитаться, если данные под ним отфильтруют или поправят."""
    import io

    from openpyxl import load_workbook

    student = enrolled('__ret_formula__')
    renewals_fixture.deal(student, world.pipeline, world.won, 1, closed_at='2026-07-15')

    content, _rows, _name = retention.build()
    ws = load_workbook(io.BytesIO(content))['Воронка по циклам']

    formulas = [
        cell.value for row in ws.iter_rows(min_col=8, max_col=8)
        for cell in row if isinstance(cell.value, str) and cell.value.startswith('=')
    ]
    assert formulas and all('IFERROR' in f for f in formulas)
