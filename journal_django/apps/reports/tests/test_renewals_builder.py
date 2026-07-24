"""Тесты построителя «Отчёта по продлениям»: выборка по РЕАЛЬНЫМ датам сделки
(outcome_at для закрытой, stage_entered_at для открытой) + рендер xlsx."""
from __future__ import annotations

import io
from datetime import datetime

import pytest
from django.utils import timezone

from apps.reports.builders import renewals

pytestmark = pytest.mark.django_db


def _dt(y, m, d, hh=12):
    """tz-aware datetime в текущей зоне (Europe/Moscow)."""
    return datetime(y, m, d, hh, tzinfo=timezone.get_current_timezone())


@pytest.fixture
def stages(renewals_fixture):
    """Общая throwaway-воронка со стадиями разных видов."""
    f = renewals_fixture
    pipe = f.pipeline()
    return {
        'f': f,
        'pipe': pipe,
        'думает': f.stage(pipe, 'thinking', 'Думает', 'decision', sort_order=5),
        'won': f.stage(pipe, 'won', 'Продлён', 'won', sort_order=8),
        'lost': f.stage(pipe, 'lost', 'Ушёл', 'lost', sort_order=9),
    }


def test_open_deal_uses_stage_entered_at(stages):
    f = stages['f']
    sid = f.student('Иванов Иван')
    aid = f.account('Петров Пётр')
    f.deal(sid, stages['pipe'], stages['думает'], cycle_no=1, assignee_id=aid,
           entered_at=_dt(2026, 5, 10))

    rows = renewals.collect_rows(2026, 5)

    assert len(rows) == 1
    r = rows[0]
    assert r['student_name'] == 'Иванов Иван'
    assert r['assignee_name'] == 'Петров Пётр'
    assert r['active_stage'] == 'Думает'
    assert r['closed_stage'] == ''
    assert r['entered_at'].isoformat() == '2026-05-10'


def test_closed_deal_uses_outcome_at(stages):
    """Закрытая сделка попадает в месяц outcome_at, дата = outcome_at (как в карточке)."""
    f = stages['f']
    sid = f.student('Сидоров Сидор')
    # встала на стадию в июле (stage_entered_at), но закрыта в августе 2025
    f.deal(sid, stages['pipe'], stages['won'], cycle_no=1,
           entered_at=_dt(2025, 7, 1), closed_at=_dt(2025, 8, 28))

    # в июле её нет — ориентируемся на дату закрытия, не на stage_entered_at
    assert renewals.collect_rows(2025, 7) == []

    rows = renewals.collect_rows(2025, 8)
    assert len(rows) == 1
    assert rows[0]['closed_stage'] == 'Продлён'
    assert rows[0]['active_stage'] == ''
    assert rows[0]['entered_at'].isoformat() == '2025-08-28'
    assert rows[0]['assignee_name'] == '—'  # без ответственного


def test_open_matured_deal_uses_due_at_over_stage_entered(stages):
    """«Ждём продление»: реальная дата — due_at (день 4-го урока), даже если
    stage_entered_at = дата пересбора. Отчёт берёт due_at."""
    f = stages['f']
    aw = f.stage(stages['pipe'], 'awaiting_renewal', 'Ждём продление', 'decision', is_auto=True, sort_order=6)
    sid = f.student('Медведев Миша')
    # цикл созрел 12.03.2026 (due_at), а на стадию «встал» по данным пересбора 20.07
    f.deal(sid, stages['pipe'], aw, cycle_no=3,
           entered_at=_dt(2026, 7, 20), due_at='2026-03-12')

    assert renewals.collect_rows(2026, 7) == []  # по дате пересбора НЕ считаем

    rows = renewals.collect_rows(2026, 3)
    assert len(rows) == 1
    assert rows[0]['active_stage'] == 'Ждём продление'
    assert rows[0]['entered_at'].isoformat() == '2026-03-12'


def test_deal_outside_month_excluded(stages):
    f = stages['f']
    sid = f.student('Волков Волк')
    f.deal(sid, stages['pipe'], stages['думает'], cycle_no=1, entered_at=_dt(2026, 4, 15))

    assert renewals.collect_rows(2026, 5) == []


def test_multiple_deals_per_student_multiple_rows(stages):
    f = stages['f']
    sid = f.student('Егоров Егор')
    f.deal(sid, stages['pipe'], stages['won'], cycle_no=1, closed_at=_dt(2026, 5, 5))
    f.deal(sid, stages['pipe'], stages['думает'], cycle_no=2, entered_at=_dt(2026, 5, 6))

    rows = renewals.collect_rows(2026, 5)

    assert len(rows) == 2
    # сортировка по cycle_no внутри ученика
    assert rows[0]['closed_stage'] == 'Продлён'
    assert rows[1]['active_stage'] == 'Думает'


def test_lost_deal_is_closed_column(stages):
    f = stages['f']
    sid = f.student('Зайцев Заяц')
    f.deal(sid, stages['pipe'], stages['lost'], cycle_no=1, closed_at=_dt(2026, 5, 20))

    rows = renewals.collect_rows(2026, 5)
    assert rows[0]['closed_stage'] == 'Ушёл'
    assert rows[0]['active_stage'] == ''


def test_render_workbook_readable_with_header(stages):
    f = stages['f']
    sid = f.student('Тестов Тест')
    f.deal(sid, stages['pipe'], stages['думает'], cycle_no=1, entered_at=_dt(2026, 5, 10))
    rows = renewals.collect_rows(2026, 5)

    content = renewals.render_workbook(rows, 2026, 5)

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    assert 'Отчёт по продлениям' in str(ws.cell(row=1, column=1).value)
    assert ws.cell(row=2, column=1).value == 'ФИО ученика'
    assert ws.cell(row=2, column=2).value == 'Цикл сделки'
    assert ws.cell(row=2, column=6).value == 'Дата переноса в эту стадию'
    assert ws.cell(row=3, column=1).value == 'Тестов Тест'
    assert ws.cell(row=3, column=2).value == 1  # cycle_no
