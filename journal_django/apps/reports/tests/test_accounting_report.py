"""Тесты бухгалтерского отчёта в разделе «Отчёты»: builder + run/download через Celery."""
from __future__ import annotations

import io

import pytest

from apps.reports import services
from apps.reports.builders import accounting
from apps.reports.models import ReportType

pytestmark = pytest.mark.django_db

BASE = '/api/admin/reports'
RUN = f'{BASE}/accounting_month/run'


def test_builder_returns_valid_workbook():
    content, count, filename = accounting.build('2026-07')
    assert filename == 'accounting_2026-07.xlsx'
    assert isinstance(count, int)
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(content)).active
    assert ws.cell(row=1, column=1).value == 'ФИО ученика'
    assert ws.cell(row=1, column=3).value == 'Посещено уроков за месяц'


def test_build_report_accounting_via_service():
    content, count, filename = services.build_report(ReportType.ACCOUNTING_MONTH, {'month': '2026-07'})
    assert filename == 'accounting_2026-07.xlsx'
    assert content[:2] == b'PK'  # zip-сигнатура xlsx


def test_run_and_download_accounting(admin_client):
    run = admin_client.post(RUN, {'month': '2026-07'}, format='json')
    assert run.status_code == 202
    task_id = run.json()['task_id']

    st = admin_client.get(f'{BASE}/status/{task_id}')
    assert st.json()['state'] == 'SUCCESS'
    assert st.json()['filename'] == 'accounting_2026-07.xlsx'

    dl = admin_client.get(f'{BASE}/download/{task_id}')
    assert dl.status_code == 200
    assert 'spreadsheetml' in dl['Content-Type']


def test_bad_month_rejected(admin_client):
    assert admin_client.post(RUN, {'month': '2026-13'}, format='json').status_code == 400


def test_future_month_rejected(admin_client):
    assert admin_client.post(RUN, {'month': '2999-01'}, format='json').status_code == 400


def test_missing_month_rejected(admin_client):
    assert admin_client.post(RUN, {}, format='json').status_code == 400
