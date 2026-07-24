"""Тесты services.build_report и celery-задачи (результат — base64, без хранения в БД)."""
from __future__ import annotations

import base64
import io
from datetime import datetime

import pytest
from django.utils import timezone

from apps.reports import services
from apps.reports.models import ReportType
from apps.reports.tasks import generate_report_task

pytestmark = pytest.mark.django_db


def _dt(y, m, d):
    return datetime(y, m, d, 12, tzinfo=timezone.get_current_timezone())


def test_build_report_renewals(renewals_fixture):
    f = renewals_fixture
    pipe = f.pipeline()
    st = f.stage(pipe, 'thinking', 'Думает', 'decision')
    sid = f.student('Иванов Иван')
    f.deal(sid, pipe, st, cycle_no=1, entered_at=_dt(2026, 5, 10))

    content, row_count, filename = services.build_report(
        ReportType.RENEWALS_MONTH, {'year': 2026, 'month': 5})

    assert row_count == 1
    assert filename == 'renewals_2026-05.xlsx'
    from openpyxl import load_workbook
    assert load_workbook(io.BytesIO(content)).active.cell(row=3, column=1).value == 'Иванов Иван'


def test_build_report_unknown_type_raises():
    with pytest.raises(services.UnknownReportType):
        services.build_report('no_such_report', {})


def test_task_returns_base64_content():
    """Задача возвращает готовый файл в результате (base64), а не пишет в БД."""
    result = generate_report_task.apply(
        args=[ReportType.RENEWALS_MONTH, {'year': 2026, 'month': 5}]).result

    assert set(result) == {'filename', 'row_count', 'content_b64'}
    raw = base64.b64decode(result['content_b64'])
    from openpyxl import load_workbook
    assert load_workbook(io.BytesIO(raw)).active.cell(row=2, column=1).value == 'ФИО ученика'
