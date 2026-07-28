"""
Тесты отчёта по посещаемости за месяц (apps.lessons.attendance_report).

Лежат в apps/reports/tests намеренно: этот пакет НЕ переопределяет
django_db_setup (см. conftest.py рядом) → pytest-django поднимает свежую
мигрированную test_journal_test, где состав учеников детерминирован. Отчёт
читает ВСЕХ учеников базы, поэтому на общей persistent journal_test проверки
вида «в отчёте ровно N строк» были бы недостоверны.

См. docs/superpowers/specs/2026-07-27-attendance-monthly-report-design.md
"""
from __future__ import annotations

import datetime
import io

import openpyxl
import pytest

from apps.directions.models import Direction
from apps.extra_lessons.models import (
    BURNED as RESOLUTION_BURNED,
)
from apps.extra_lessons.models import (
    EXTRA as RESOLUTION_EXTRA,
)
from apps.extra_lessons.models import (
    MAKEUP,
    MAKEUP_DONE,
    PENDING,
    AbsenceResolution,
)
from apps.groups.models import Group
from apps.lessons.attendance_report import (
    ABSENT,
    BURNED,
    EMPTY,
    EXTRA_OVER_COURSE,
    KIND_ABSENT,
    KIND_BURNED,
    KIND_MADE_UP,
    KIND_PRESENT,
    MADE_UP,
    NO_GROUP,
    PRESENT,
    collect_month,
    render_bytes,
)
from apps.lessons.models import Lesson, LessonAttendance
from apps.memberships.models import GroupMembership
from apps.students.models import Student
from apps.teachers.models import Teacher

pytestmark = pytest.mark.django_db


@pytest.fixture
def data():
    """Фабрика доменных объектов отчёта (ORM, свежая test-БД чистится сама)."""

    class F:
        def __init__(self):
            self.teacher = Teacher.objects.create(
                name='__att_teacher__',
                created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )
            self.direction = Direction.objects.create(
                name='__att_dir__', total_lessons=8, active=True,
            )

        def group(self, name: str) -> Group:
            return Group.objects.create(
                name=name, direction=self.direction, teacher=self.teacher,
                is_individual=False, created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )

        def student(self, full_name: str) -> Student:
            return Student.objects.create(
                full_name=full_name,
                created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )

        def membership(self, student: Student, group: Group, active: bool = True):
            return GroupMembership.objects.create(
                student=student, group=group, active=active,
            )

        def lesson(self, group: Group, date: str, number: float,
                   lesson_type: str = 'regular') -> Lesson:
            # submitted_by_token разводит уроки по типу: у пропуска и его
            # отработки совпадают дата+группа+номер (отработка в день пропуска —
            # реальный случай), а natural key требует различия.
            return Lesson.objects.create(
                group=group, teacher=self.teacher, lesson_date=date,
                lesson_number=number, lesson_duration_minutes=90,
                lesson_type=lesson_type,
                submitted_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
                submitted_by_token=f'__att_test_{lesson_type}__',
            )

        def attend(self, lesson: Lesson, student: Student, present: bool,
                   is_free: bool = False, unpaid_skip: bool = False):
            return LessonAttendance.objects.create(
                lesson=lesson, student=student, present=present,
                is_free=is_free, unpaid_skip=unpaid_skip,
            )

        def makeup(self, student: Student, missed: Lesson, fact: Lesson | None = None,
                   status: str = MAKEUP_DONE) -> AbsenceResolution:
            """Резолюция пропуска: отработка (fact.lesson_type='extra') или
            сгорание (fact.lesson_type='burned', status=RESOLUTION_BURNED)."""
            return AbsenceResolution.objects.create(
                missed_lesson=missed, student=student, kind=MAKEUP,
                status=status, fact_lesson=fact,
                scheduled_date=fact.lesson_date if fact else None,
            )

        def extra_over_course(self, student: Student, group: Group,
                              fact: Lesson) -> AbsenceResolution:
            """Доп.урок СВЕРХ курса: пропуска нет (missed_lesson=NULL)."""
            return AbsenceResolution.objects.create(
                missed_lesson=None, group=group, student=student,
                kind=RESOLUTION_EXTRA, status=MAKEUP_DONE, fact_lesson=fact,
                scheduled_date=fact.lesson_date,
            )

    return F()


def _row(rows, full_name, group_name=None):
    """Единственная строка отчёта по ученику (+группе), иначе AssertionError."""
    found = [r for r in rows
             if r.full_name == full_name
             and (group_name is None or r.group_label == group_name)]
    assert len(found) == 1, f'ожидалась 1 строка, найдено {len(found)}'
    return found[0]


def test_statuses_present_absent(data):
    """Был / Не был; нерешённый пропуск остаётся «Не был»."""
    g = data.group('__att_g1__')
    s = data.student('__att_s1__')
    data.membership(s, g)
    l1 = data.lesson(g, '2026-07-07', 1)
    l2 = data.lesson(g, '2026-07-14', 2)
    data.attend(l1, s, present=True)
    data.attend(l2, s, present=False)
    data.makeup(s, missed=l2, fact=None, status=PENDING)

    row = _row(collect_month('2026-07'), '__att_s1__', '__att_g1__')

    assert [(c.status, c.kind) for c in row.cells] == [
        (PRESENT, KIND_PRESENT), (ABSENT, KIND_ABSENT),
    ]
    assert [c.date for c in row.cells] == [
        datetime.date(2026, 7, 7), datetime.date(2026, 7, 14),
    ]
    assert (row.present_count, row.absent_count, row.made_up_count) == (1, 1, 0)


def test_makeup_collapses_onto_the_missed_lesson(data):
    """Отработка в том же месяце НЕ создаёт колонку — ложится на пропуск."""
    g = data.group('__att_g2__')
    s = data.student('__att_s2__')
    data.membership(s, g)
    missed = data.lesson(g, '2026-07-19', 2)
    fact = data.lesson(g, '2026-07-19', 2, lesson_type='extra')
    data.attend(missed, s, present=False)
    data.attend(fact, s, present=True)
    data.makeup(s, missed=missed, fact=fact)
    data.attend(data.lesson(g, '2026-07-26', 3), s, present=True)

    row = _row(collect_month('2026-07'), '__att_s2__', '__att_g2__')

    # Ровно две колонки: урок 2 (отработан) и урок 3. Никакого сдвига.
    assert [(c.date, c.status) for c in row.cells] == [
        (datetime.date(2026, 7, 19), MADE_UP),
        (datetime.date(2026, 7, 26), PRESENT),
    ]
    assert (row.present_count, row.absent_count, row.made_up_count) == (1, 0, 1)


def test_burn_collapses_onto_the_missed_lesson(data):
    """Сгорание в том же месяце тоже ложится на пропуск, своей колонки нет."""
    g = data.group('__att_g3__')
    s = data.student('__att_s3__')
    data.membership(s, g)
    missed = data.lesson(g, '2026-07-11', 1)
    fact = data.lesson(g, '2026-07-23', 1, lesson_type='burned')
    data.attend(missed, s, present=False)
    data.attend(fact, s, present=True)   # запись сгорания в БД идёт present=true
    data.makeup(s, missed=missed, fact=fact, status=RESOLUTION_BURNED)

    row = _row(collect_month('2026-07'), '__att_s3__', '__att_g3__')

    assert [(c.date, c.status, c.kind) for c in row.cells] == [
        (datetime.date(2026, 7, 11), BURNED, KIND_BURNED),
    ]
    # «Сгорел» не входит ни в один счётчик.
    assert (row.present_count, row.absent_count, row.made_up_count) == (0, 0, 0)


def test_makeup_for_previous_month_gets_its_own_column(data):
    """Пропуск в июне, отработка в июле: ложиться не на что → своя колонка."""
    g = data.group('__att_g4__')
    s = data.student('__att_s4__')
    data.membership(s, g)
    missed = data.lesson(g, '2026-06-14', 1)
    fact = data.lesson(g, '2026-07-19', 1, lesson_type='extra')
    data.attend(missed, s, present=False)
    data.attend(fact, s, present=True)
    data.makeup(s, missed=missed, fact=fact)

    july = _row(collect_month('2026-07'), '__att_s4__', '__att_g4__')
    june = _row(collect_month('2026-06'), '__att_s4__', '__att_g4__')

    assert [(c.date, c.status, c.kind) for c in july.cells] == [
        (datetime.date(2026, 7, 19), 'Отработка за 14.06.2026', KIND_MADE_UP),
    ]
    # В июне пропуск остаётся пропуском: закрыт он в июле и там же показан —
    # одно событие не должно попасть в два месячных отчёта.
    assert [(c.date, c.status) for c in june.cells] == [
        (datetime.date(2026, 6, 14), ABSENT),
    ]


def test_burn_for_previous_month_gets_its_own_column(data):
    """Сгорание за пропуск прошлого месяца — своя колонка с пояснением."""
    g = data.group('__att_g5__')
    s = data.student('__att_s5__')
    data.membership(s, g)
    missed = data.lesson(g, '2026-02-06', 1)
    fact = data.lesson(g, '2026-07-23', 1, lesson_type='burned')
    data.attend(missed, s, present=False)
    data.attend(fact, s, present=True)
    data.makeup(s, missed=missed, fact=fact, status=RESOLUTION_BURNED)

    row = _row(collect_month('2026-07'), '__att_s5__', '__att_g5__')

    assert [(c.status, c.kind) for c in row.cells] == [
        ('Сгорел (за 06.02.2026)', KIND_BURNED),
    ]


def test_extra_over_course_keeps_own_column(data):
    """Доп.урок СВЕРХ курса не привязан к пропуску → отдельная колонка."""
    g = data.group('__att_g6__')
    s = data.student('__att_s6__')
    data.membership(s, g)
    data.attend(data.lesson(g, '2026-07-07', 1), s, present=True)
    fact = data.lesson(g, '2026-07-24', 2, lesson_type='extra')
    data.attend(fact, s, present=True)
    data.extra_over_course(s, group=g, fact=fact)

    row = _row(collect_month('2026-07'), '__att_s6__', '__att_g6__')

    assert [(c.status, c.kind) for c in row.cells] == [
        (PRESENT, KIND_PRESENT), (EXTRA_OVER_COURSE, KIND_PRESENT),
    ]
    assert (row.present_count, row.absent_count, row.made_up_count) == (2, 0, 0)


def test_free_is_present_and_unpaid_skip_is_absent(data):
    """is_free схлопывается в «Был», unpaid_skip — в «Не был»."""
    g = data.group('__att_g7__')
    s = data.student('__att_s7__')
    data.membership(s, g)
    l1 = data.lesson(g, '2026-07-07', 1)
    l2 = data.lesson(g, '2026-07-14', 2)
    data.attend(l1, s, present=True, is_free=True)
    data.attend(l2, s, present=False, unpaid_skip=True)

    row = _row(collect_month('2026-07'), '__att_s7__', '__att_g7__')

    assert [c.status for c in row.cells] == [PRESENT, ABSENT]


def test_student_in_two_groups_gets_two_rows(data):
    """Ученик в двух группах = две строки, даты не смешиваются."""
    ga = data.group('__att_A__')
    gb = data.group('__att_B__')
    s = data.student('__att_s8__')
    data.membership(s, ga)
    data.membership(s, gb)
    data.attend(data.lesson(ga, '2026-07-07', 1), s, present=True)
    data.attend(data.lesson(gb, '2026-07-08', 1), s, present=False)

    rows = collect_month('2026-07')

    assert [c.status for c in _row(rows, '__att_s8__', '__att_A__').cells] == [PRESENT]
    assert [c.status for c in _row(rows, '__att_s8__', '__att_B__').cells] == [ABSENT]


def test_all_students_present_even_without_lessons(data):
    """Требование руководства: в отчёте все ученики базы."""
    g = data.group('__att_g9__')
    idle = data.student('__att_idle__')          # членство есть, уроков в месяце нет
    data.membership(idle, g)
    orphan = data.student('__att_orphan__')      # ни группы, ни уроков

    rows = collect_month('2026-07')

    assert _row(rows, '__att_idle__', '__att_g9__').cells == []
    assert _row(rows, '__att_orphan__').group_label == NO_GROUP
    assert {r.student_id for r in rows} == set(Student.objects.values_list('id', flat=True))
    assert orphan.id in {r.student_id for r in rows}


def test_half_lessons_counted_as_sessions_not_course_lessons(data):
    """45-минутные занятия: каждое — своя колонка, итог считает ЗАНЯТИЯ.

    Half-lesson (45 мин → 0.5 урока) здесь намеренно НЕ применяется, хотя
    бухотчёт те же два занятия считает за 1 урок. Решение 2026-07-27: отчёт
    отвечает «сколько раз пришёл», а не «сколько уроков курса отработано».
    Тест держит это решение, чтобы его не «починили» заодно.
    """
    g = data.group('__att_g45__')
    s = data.student('__att_s45__')
    data.membership(s, g)
    first = data.lesson(g, '2026-07-24', 37.5)
    second = data.lesson(g, '2026-07-24', 38)
    for lesson in (first, second):
        lesson.lesson_duration_minutes = 45
        lesson.save(update_fields=['lesson_duration_minutes'])
    data.attend(first, s, present=True)
    data.attend(second, s, present=True)

    row = _row(collect_month('2026-07'), '__att_s45__', '__att_g45__')

    # Две колонки с одной датой, не одна схлопнутая: статусы независимы.
    assert [(c.date, c.status) for c in row.cells] == [
        (datetime.date(2026, 7, 24), PRESENT),
        (datetime.date(2026, 7, 24), PRESENT),
    ]
    assert (row.present_count, row.absent_count, row.made_up_count) == (2, 0, 0)


def test_month_boundaries(data):
    """Уроки соседних месяцев в отчёт за июль не попадают."""
    g = data.group('__att_g10__')
    s = data.student('__att_s10__')
    data.membership(s, g)
    data.attend(data.lesson(g, '2026-06-30', 1), s, present=True)
    data.attend(data.lesson(g, '2026-07-01', 2), s, present=True)
    data.attend(data.lesson(g, '2026-07-31', 3), s, present=False)
    data.attend(data.lesson(g, '2026-08-01', 4), s, present=True)

    row = _row(collect_month('2026-07'), '__att_s10__', '__att_g10__')

    assert [c.date for c in row.cells] == [
        datetime.date(2026, 7, 1), datetime.date(2026, 7, 31),
    ]


def test_invalid_month_raises(data):
    with pytest.raises(ValueError):
        collect_month('2026-13')


def test_render_layout(data):
    """Рендер: пара строк на ученика+группу, merge ФИО/группы, прочерки справа."""
    ga = data.group('__att_r1__')
    gb = data.group('__att_r2__')
    s_many = data.student('__att_r_many__')
    s_few = data.student('__att_r_few__')
    data.membership(s_many, ga)
    data.membership(s_few, gb)
    for day, number in ((7, 1), (14, 2), (21, 3)):
        data.attend(data.lesson(ga, f'2026-07-{day:02d}', number), s_many, present=True)
    data.attend(data.lesson(gb, '2026-07-08', 1), s_few, present=False)

    rows = collect_month('2026-07')
    ws = openpyxl.load_workbook(io.BytesIO(render_bytes(rows))).active

    # 1 шапка + по 2 строки на каждую пару ученик+группа.
    assert ws.max_row == 1 + 2 * len(rows)
    # Колонок = 2 фиксированные + максимум уроков (3) + 3 итоговые.
    assert ws.max_column == 2 + 3 + 3
    assert [c.value for c in ws[1]] == [
        'ФИО ученика', 'Группа', 'Урок 1', 'Урок 2', 'Урок 3',
        'Итого «Был»', 'Итого «Не был»', 'Итого «Отработано»',
    ]

    index = {(r.full_name, r.group_label): i for i, r in enumerate(rows)}
    top = 2 + 2 * index[('__att_r_few__', '__att_r2__')]
    assert ws.cell(row=top, column=1).value == '__att_r_few__'
    assert ws.cell(row=top, column=2).value == '__att_r2__'
    date_cell = ws.cell(row=top, column=3)
    # openpyxl отдаёт дату обратно как datetime — важно, что это НЕ строка:
    # в Excel ячейка остаётся датой и форматируется по DD.MM.YYYY.
    assert date_cell.value == datetime.datetime(2026, 7, 8)
    assert date_cell.number_format == 'DD.MM.YYYY'
    assert ws.cell(row=top + 1, column=3).value == ABSENT
    # У этой пары всего один урок — остальные колонки прочерк в обеих строках.
    assert ws.cell(row=top, column=4).value == EMPTY
    assert ws.cell(row=top + 1, column=4).value == EMPTY
    assert [ws.cell(row=top, column=c).value for c in (6, 7, 8)] == [0, 1, 0]
    # ФИО/группа/итоги объединены по вертикали на обе строки пары.
    merged = {str(m) for m in ws.merged_cells.ranges}
    assert {f'A{top}:A{top + 1}', f'B{top}:B{top + 1}', f'F{top}:F{top + 1}',
            f'G{top}:G{top + 1}', f'H{top}:H{top + 1}'} <= merged
