"""
Тесты прогноза отработки денег на данных (apps.finances.revenue_forecast).

Лежат в apps/reports/tests намеренно: пакет НЕ переопределяет django_db_setup
(см. conftest.py рядом) → свежая мигрированная test_journal_test. Прогноз идёт
через fifo_inputs(), который читает ВСЕ оплаты системы, поэтому на общей
persistent journal_test результат был бы недетерминирован.

Арифметика нарезки покрыта отдельно и без БД:
apps/finances/tests/test_revenue_forecast_split.py

См. docs/superpowers/specs/2026-07-27-revenue-forecast-design.md
"""
from __future__ import annotations

import datetime
import io
from decimal import Decimal

import openpyxl
import pytest

from apps.directions.models import Direction
from apps.finances.revenue_forecast import (
    NO_DIRECTION,
    collect_forecast,
    month_label,
    render_bytes,
)
from apps.groups.models import Group
from apps.lessons.models import Lesson, LessonAttendance
from apps.payments.models import Payment
from apps.students.models import Student
from apps.teachers.models import Teacher

pytestmark = pytest.mark.django_db

MONTH = '2026-07'


def _D(x):
    return Decimal(str(x))


@pytest.fixture
def data():
    """Фабрика оплат/уроков для прогноза."""

    class F:
        def __init__(self):
            self.teacher = Teacher.objects.create(
                name='__fc_teacher__',
                created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )

        def direction(self, name: str) -> Direction:
            return Direction.objects.create(name=name, total_lessons=48, active=True)

        def group(self, name: str, direction: Direction) -> Group:
            return Group.objects.create(
                name=name, direction=direction, teacher=self.teacher,
                is_individual=False,
                created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )

        def student(self, full_name: str) -> Student:
            return Student.objects.create(
                full_name=full_name,
                created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )

        def payment(self, student: Student, direction: Direction | None,
                    subscriptions: int, total: str, paid_at: str,
                    kind: str = 'purchase') -> Payment:
            """1 абонемент = 4 урока — инвариант данных проекта."""
            lessons = subscriptions * 4
            return Payment.objects.create(
                student=student, direction=direction,
                subscriptions_count=subscriptions, lessons_count=lessons,
                kind=kind, unit_price=_D(total) / _D(subscriptions),
                total_amount=_D(total), paid_at=paid_at,
                created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
            )

        def attend(self, student: Student, group: Group, date: str, number: float,
                   duration: int = 90):
            lesson = Lesson.objects.create(
                group=group, teacher=self.teacher, lesson_date=date,
                lesson_number=number, lesson_duration_minutes=duration,
                lesson_type='regular',
                submitted_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
                submitted_by_token=f'__fc_{date}_{number}__',
            )
            LessonAttendance.objects.create(lesson=lesson, student=student, present=True)
            return lesson

    return F()


def _row(forecast, full_name, direction_name=None):
    found = [r for r in forecast.rows
             if r.full_name == full_name
             and (direction_name is None or r.direction_name == direction_name)]
    assert len(found) == 1, f'ожидалась 1 строка, найдено {len(found)}'
    return found[0]


def test_payment_spreads_one_subscription_per_month(data):
    """9 абонементов за 50 000 → 9 месяцев подряд, сумма сходится в копейку."""
    d = data.direction('__fc_dir__')
    s = data.student('__fc_s1__')
    data.payment(s, d, subscriptions=9, total='50000.00', paid_at='2026-07-10')

    row = _row(collect_forecast(MONTH), '__fc_s1__', '__fc_dir__')

    assert len(row.by_month) == 9
    assert list(row.by_month)[0] == '2026-07'
    assert list(row.by_month)[-1] == '2027-03'
    assert row.remaining_lessons == _D(36)
    assert row.remaining_value == _D('50000.00')
    assert sum(row.by_month.values(), _D(0)) == _D('50000.00')
    # 50000 / 9 = 5555.5555…: помесячно округляем, невязку отдаём последнему месяцу.
    assert row.by_month['2026-07'] == _D('5555.56')
    assert row.by_month['2027-03'] == _D('5555.52')


def test_worked_off_lessons_are_excluded(data):
    """Прогноз — только НЕотработанный остаток: посещённые уроки уходят из хвоста."""
    d = data.direction('__fc_dir2__')
    g = data.group('__fc_g2__', d)
    s = data.student('__fc_s2__')
    data.payment(s, d, subscriptions=2, total='8000.00', paid_at='2026-05-05')
    for i in range(1, 5):                      # 4 урока из 8 уже отработаны
        data.attend(s, g, f'2026-05-{10 + i:02d}', i)

    row = _row(collect_forecast(MONTH), '__fc_s2__', '__fc_dir2__')

    assert row.remaining_lessons == _D(4)
    assert row.by_month == {'2026-07': _D('4000.00')}


def test_layout_starts_from_report_month_not_payment_month(data):
    """Раскладка идёт от месяца отчёта: старая неотработанная оплата съезжает в хвост."""
    d = data.direction('__fc_dir3__')
    s = data.student('__fc_s3__')
    data.payment(s, d, subscriptions=2, total='8000.00', paid_at='2026-02-01')

    row = _row(collect_forecast(MONTH), '__fc_s3__', '__fc_dir3__')

    # Оплата февральская, но месяцы прогноза — июль и август.
    assert list(row.by_month) == ['2026-07', '2026-08']


def test_two_directions_are_separate_rows_and_run_in_parallel(data):
    """Курсы идут параллельно: у каждого направления свои 4 урока в месяц."""
    minecraft = data.direction('__fc_minecraft__')
    blender = data.direction('__fc_blender__')
    s = data.student('__fc_s4__')
    data.payment(s, minecraft, subscriptions=3, total='30000.00', paid_at='2026-07-01')
    data.payment(s, blender, subscriptions=1, total='5000.00', paid_at='2026-07-02')

    forecast = collect_forecast(MONTH)

    mine = _row(forecast, '__fc_s4__', '__fc_minecraft__')
    blend = _row(forecast, '__fc_s4__', '__fc_blender__')
    # Оба стартуют с июля, а не встают в очередь друг за другом.
    assert list(mine.by_month) == ['2026-07', '2026-08', '2026-09']
    assert list(blend.by_month) == ['2026-07']


def test_refund_shrinks_the_tail(data):
    """Возврат гасит остаток — прогнозировать возвращённые деньги нечего."""
    d = data.direction('__fc_dir5__')
    s = data.student('__fc_s5__')
    data.payment(s, d, subscriptions=3, total='12000.00', paid_at='2026-06-01')
    Payment.objects.create(
        student=s, direction=d, subscriptions_count=-1, lessons_count=-4,
        kind='refund', unit_price=_D('4000.00'), total_amount=_D('-4000.00'),
        paid_at='2026-06-20',
        created_at=datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC),
    )

    row = _row(collect_forecast(MONTH), '__fc_s5__', '__fc_dir5__')

    assert row.remaining_lessons == _D(8)
    assert row.remaining_value == _D('8000.00')


def test_student_without_remaining_money_is_absent(data):
    """Ученик, отработавший всё, в прогноз не попадает — прогнозировать нечего."""
    d = data.direction('__fc_dir6__')
    g = data.group('__fc_g6__', d)
    s = data.student('__fc_s6__')
    data.payment(s, d, subscriptions=1, total='4000.00', paid_at='2026-05-01')
    for i in range(1, 5):
        data.attend(s, g, f'2026-05-{10 + i:02d}', i)

    assert [r for r in collect_forecast(MONTH).rows if r.full_name == '__fc_s6__'] == []


def test_legacy_payment_without_direction(data):
    """Легаси-оплата без направления попадает в лист «Без направления»."""
    s = data.student('__fc_s7__')
    data.payment(s, None, subscriptions=1, total='4000.00', paid_at='2026-07-01')

    row = _row(collect_forecast(MONTH), '__fc_s7__')

    assert row.direction_id is None
    assert row.direction_name == NO_DIRECTION


def test_half_lesson_leaves_fractional_tail(data):
    """45-минутные занятия: остаток дробный, деньги считаются в уроках."""
    d = data.direction('__fc_dir8__')
    g = data.group('__fc_g8__', d)
    s = data.student('__fc_s8__')
    data.payment(s, d, subscriptions=1, total='4000.00', paid_at='2026-05-01')
    data.attend(s, g, '2026-05-11', 0.5, duration=45)     # 0.5 урока

    row = _row(collect_forecast(MONTH), '__fc_s8__', '__fc_dir8__')

    assert row.remaining_lessons == _D('3.5')
    assert row.by_month == {'2026-07': _D('3500.00')}


def test_months_are_a_continuous_range(data):
    """Колонки — сплошной ряд месяцев, даже если у кого-то в середине пусто."""
    d = data.direction('__fc_dir9__')
    long_payer = data.student('__fc_long__')
    short_payer = data.student('__fc_short__')
    data.payment(long_payer, d, subscriptions=4, total='16000.00', paid_at='2026-07-01')
    data.payment(short_payer, d, subscriptions=1, total='4000.00', paid_at='2026-07-01')

    forecast = collect_forecast(MONTH)

    assert forecast.months == ['2026-07', '2026-08', '2026-09', '2026-10']
    assert list(_row(forecast, '__fc_short__').by_month) == ['2026-07']


def test_render_sheet_per_direction(data):
    """Рендер: лист «Сводка» + лист на направление, строка ИТОГО, месяцы-колонки."""
    a = data.direction('__fc_A__')
    b = data.direction('__fc_B__')
    s1 = data.student('__fc_r1__')
    s2 = data.student('__fc_r2__')
    data.payment(s1, a, subscriptions=2, total='8000.00', paid_at='2026-07-01')
    data.payment(s2, a, subscriptions=1, total='5000.00', paid_at='2026-07-01')
    data.payment(s1, b, subscriptions=1, total='3000.00', paid_at='2026-07-01')

    forecast = collect_forecast(MONTH)
    wb = openpyxl.load_workbook(io.BytesIO(render_bytes(forecast)))

    assert wb.sheetnames[0] == 'Сводка'
    assert {'__fc_A__', '__fc_B__'} <= set(wb.sheetnames)

    ws = wb['__fc_A__']
    assert ws.cell(row=1, column=1).value == 'ФИО ученика'
    assert ws.cell(row=1, column=5).value == month_label(forecast.months[0])
    rows_by_name = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    assert ws.cell(row=rows_by_name['__fc_r1__'], column=5).value == 4000.0
    assert ws.cell(row=rows_by_name['__fc_r2__'], column=5).value == 5000.0
    # Последняя строка листа — ИТОГО по направлению.
    assert ws.cell(row=ws.max_row, column=1).value == 'ИТОГО'
    assert ws.cell(row=ws.max_row, column=3).value == 13000.0

    summary = wb['Сводка']
    summary_rows = {summary.cell(row=r, column=1).value: r
                    for r in range(2, summary.max_row + 1)}
    assert summary.cell(row=summary_rows['__fc_A__'], column=4).value == 13000.0
    assert summary.cell(row=summary_rows['__fc_B__'], column=4).value == 3000.0
    assert summary.cell(row=summary.max_row, column=1).value == 'ИТОГО'
    assert summary.cell(row=summary.max_row, column=4).value == 16000.0


def test_invalid_month_raises(data):
    with pytest.raises(ValueError):
        collect_forecast('2026-13')


# ---------------------------------------------------------------------------
# Режим полной истории
# ---------------------------------------------------------------------------

def test_full_history_shows_past_as_fact_and_future_as_plan(data):
    """Прошлые месяцы — факт отработки, будущие — прогноз, в одной строке."""
    d = data.direction('__fh_dir1__')
    g = data.group('__fh_g1__', d)
    s = data.student('__fh_s1__')
    data.payment(s, d, subscriptions=3, total='12000.00', paid_at='2026-05-01')
    for i in range(1, 5):                       # 4 урока отработаны в мае
        data.attend(s, g, f'2026-05-{10 + i:02d}', i)

    row = _row(collect_forecast(MONTH, full_history=True), '__fh_s1__', '__fh_dir1__')

    assert row.by_month == {
        '2026-05': _D('4000.00'),               # факт
        '2026-07': _D('4000.00'),               # план
        '2026-08': _D('4000.00'),
    }
    assert row.fact_months == {'2026-05'}
    assert row.worked_off_value == _D('4000.00')
    assert row.remaining_value == _D('8000.00')
    # Вся история строки = оплаченное по этому направлению.
    assert row.worked_off_value + row.remaining_value == _D('12000.00')


def test_current_month_fact_and_plan_do_not_double_count(data):
    """Июль уже частично отработан — план добирает лишь остаток до 4 уроков."""
    d = data.direction('__fh_dir2__')
    g = data.group('__fh_g2__', d)
    s = data.student('__fh_s2__')
    data.payment(s, d, subscriptions=2, total='8000.00', paid_at='2026-07-01')
    data.attend(s, g, '2026-07-05', 1)          # 2 урока уже отработаны в июле
    data.attend(s, g, '2026-07-12', 2)

    row = _row(collect_forecast(MONTH, full_history=True), '__fh_s2__', '__fh_dir2__')

    # Июль: 2 урока факта + 2 урока добора = 4 урока = 4000, а не 2000 + 4000.
    assert row.by_month['2026-07'] == _D('4000.00')
    assert row.by_month['2026-08'] == _D('4000.00')
    assert sum(row.by_month.values(), _D(0)) == _D('8000.00')
    assert row.worked_off_value == _D('2000.00')
    assert row.remaining_value == _D('6000.00')


def test_plain_mode_also_tops_up_the_start_month(data):
    """Тот же добор действует и в обычном режиме — без него июль задваивался."""
    d = data.direction('__fh_dir3__')
    g = data.group('__fh_g3__', d)
    s = data.student('__fh_s3__')
    data.payment(s, d, subscriptions=2, total='8000.00', paid_at='2026-07-01')
    for i in range(1, 5):                       # весь июльский абонемент отработан
        data.attend(s, g, f'2026-07-{i:02d}', i)

    row = _row(collect_forecast(MONTH), '__fh_s3__', '__fh_dir3__')

    # Июля в прогнозе нет вовсе: его 4 урока уже закрыты фактом.
    assert row.by_month == {'2026-08': _D('4000.00')}


def test_fully_worked_off_student_appears_only_in_full_history(data):
    """Ученик без остатка: в прогнозе его нет, в истории — есть."""
    d = data.direction('__fh_dir4__')
    g = data.group('__fh_g4__', d)
    s = data.student('__fh_s4__')
    data.payment(s, d, subscriptions=1, total='4000.00', paid_at='2026-04-01')
    for i in range(1, 5):
        data.attend(s, g, f'2026-04-{10 + i:02d}', i)

    assert [r for r in collect_forecast(MONTH).rows if r.full_name == '__fh_s4__'] == []

    row = _row(collect_forecast(MONTH, full_history=True), '__fh_s4__', '__fh_dir4__')
    assert row.by_month == {'2026-04': _D('4000.00')}
    assert row.remaining_value == _D('0')
    assert row.worked_off_value == _D('4000.00')


def test_fact_goes_to_lesson_direction_not_payment_direction(data):
    """Заплатил за один курс, занимался на другом: факт — на листе курса УРОКА."""
    paid_dir = data.direction('__ld_paid__')
    lesson_dir = data.direction('__ld_lesson__')
    g = data.group('__ld_g__', lesson_dir)
    s = data.student('__ld_s1__')
    # 2 абонемента = 8 уроков за 8000; 4 из них отработаны на ЧУЖОМ направлении.
    data.payment(s, paid_dir, subscriptions=2, total='8000.00', paid_at='2026-05-01')
    for i in range(1, 5):
        data.attend(s, g, f'2026-05-{10 + i:02d}', i)

    forecast = collect_forecast(MONTH, full_history=True)

    # Лист курса, на котором реально занимались: только факт, аванса тут нет.
    fact_row = _row(forecast, '__ld_s1__', '__ld_lesson__')
    assert fact_row.by_month == {'2026-05': _D('4000.00')}
    assert fact_row.worked_off_value == _D('4000.00')
    assert fact_row.remaining_value == _D('0')

    # Лист курса, за который платили: остаток аванса, но факта нет.
    plan_row = _row(forecast, '__ld_s1__', '__ld_paid__')
    assert plan_row.worked_off_value == _D('0')
    assert plan_row.by_month == {'2026-07': _D('4000.00')}
    assert plan_row.remaining_value == _D('4000.00')


def test_start_month_top_up_still_uses_payment_direction(data):
    """Добор стартового месяца считает ДЕНЬГИ, поэтому остаётся на разрезе оплаты."""
    paid_dir = data.direction('__ld_paid2__')
    lesson_dir = data.direction('__ld_lesson2__')
    g = data.group('__ld_g2__', lesson_dir)
    s = data.student('__ld_s2__')
    # 2 абонемента = 8 уроков за 8000; 2 урока чужого курса отработаны в СТАРТОВОМ месяце.
    data.payment(s, paid_dir, subscriptions=2, total='8000.00', paid_at='2026-07-01')
    data.attend(s, g, '2026-07-05', 1)
    data.attend(s, g, '2026-07-12', 2)

    forecast = collect_forecast(MONTH, full_history=True)

    # Деньги направления оплаты в июле уже потрачены на 2 урока → план добирает 2,
    # а не полный абонемент. Иначе июль завысился бы на 2000.
    plan_row = _row(forecast, '__ld_s2__', '__ld_paid2__')
    assert plan_row.by_month == {'2026-07': _D('2000.00'), '2026-08': _D('4000.00')}
    assert plan_row.remaining_value == _D('6000.00')
    assert plan_row.worked_off_value == _D('0')

    # А сам факт этих 2 уроков показан на курсе, где занимались.
    fact_row = _row(forecast, '__ld_s2__', '__ld_lesson2__')
    assert fact_row.by_month == {'2026-07': _D('2000.00')}
    assert fact_row.worked_off_value == _D('2000.00')


def test_month_kind_labels():
    forecast = collect_forecast(MONTH, full_history=True)

    assert forecast.month_kind('2026-05') == 'факт'
    assert forecast.month_kind('2026-07') == 'факт+план'
    assert forecast.month_kind('2026-09') == 'план'
    # Без истории все колонки плановые.
    assert collect_forecast(MONTH).month_kind('2026-05') == 'план'


def test_render_full_history_marks_fact_and_plan(data):
    """Рендер истории: колонка признанной выручки и подписи «· факт» / «· план»."""
    d = data.direction('__fh_R__')
    g = data.group('__fh_gR__', d)
    s = data.student('__fh_r1__')
    # 3 абонемента = 12 уроков: 4 отработаны в мае, остаток 8 → июль и август.
    data.payment(s, d, subscriptions=3, total='12000.00', paid_at='2026-05-01')
    for i in range(1, 5):
        data.attend(s, g, f'2026-05-{10 + i:02d}', i)

    forecast = collect_forecast(MONTH, full_history=True)
    ws = openpyxl.load_workbook(io.BytesIO(render_bytes(forecast)))['__fh_R__']

    header = [c.value for c in ws[1]]
    assert header[:5] == ['ФИО ученика', 'Остаток уроков', 'Остаток аванса, ₽',
                          'Цена урока, ₽', 'Признано выручки на курсе, ₽']
    assert 'Май 2026 · факт' in header
    assert 'Июль 2026 · факт+план' in header
    assert 'Август 2026 · план' in header

    row_idx = next(r for r in range(2, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == '__fh_r1__')
    assert ws.cell(row=row_idx, column=5).value == 4000.0       # признано
    assert ws.cell(row=row_idx, column=3).value == 8000.0       # остаток аванса
    may_col = header.index('Май 2026 · факт') + 1
    assert ws.cell(row=row_idx, column=may_col).value == 4000.0
