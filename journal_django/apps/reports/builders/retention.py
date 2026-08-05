"""
Построитель «Отчёта по переходимости» — на каком цикле школа теряет детей.

Параметров нет: отчёт общий, по всей истории. Период не выбирается сознательно —
вопрос «где провал» требует всей воронки целиком, а окно её обрезало бы.

ГЛАВНОЕ ОТКРЫТИЕ, НА КОТОРОМ ПОСТРОЕН ОТЧЁТ. Переходимость нельзя мерить
отметкой «Ушёл»: по всей истории школы таких отметок два десятка, потому что
ушедшего ученика обычно просто перестают вести, а его сделка остаётся висеть
в «Ждём продление». Мерить надо ФАКТОМ ДОХОДА ДО СЛЕДУЮЩЕГО ЦИКЛА: движок
продлений создаёт сделку цикла N+1 ровно тогда, когда сделка цикла N закрыта
как «Продлён» (проверено на данных: won(N) == всего(N+1) на каждом цикле).
Поэтому воронка читается по числу сделок на каждом цикле, а зависшая сделка
считается потерей, а не «в работе».

Цикл = один оплаченный абонемент = 4 урока (`apps.renewals.cycle
.LESSONS_PER_CYCLE`). Отсюда колонка «уроков курса»: цикл 9 — это уроки 33–36,
то есть конец стандартного 36-урочного курса, и провал именно там означает,
что дети не продлеваются на второй курс, а не «устают в середине».

Структура книги:
  1. «Воронка по циклам»     главный лист: где именно теряем, + графики;
  2. «Циклы × преподаватели» переход по диапазонам циклов, строка = преподаватель;
  3. «Циклы × направления»   то же, строка = направление;
  4. «Зависшие сделки»       рабочий список: с кем разобраться прямо сейчас;
  5. «Детализация»           строка = сделка, чтобы разложить любую цифру.

ОГОВОРКИ (продублированы в шапке листов, без них цифры читаются неверно):

1. Сумма строк в разрезах БОЛЬШЕ школьного итога. `renewal_deal` привязана к
   ученику и циклу — направления или группы в ней нет вовсе. Ученика связываем
   с преподавателями и направлениями через его группы, поэтому занимающийся у
   двоих попадает в обе строки. Колонку нельзя складывать.
2. «В работе» исключены из знаменателя перехода: ученик, у которого цикл идёт
   прямо сейчас, ещё не решил — считать его потерей значило бы занижать
   переход тем сильнее, чем больше у преподавателя активных учеников.
3. Служебная запись «Архив (импорт истории)» (`teachers.is_service`) идёт
   отдельной строкой и НЕ входит в итог: на ней ~80 % всех сделок школы.
"""
from __future__ import annotations

import io
from collections import defaultdict

from django.db import connection

from apps.renewals.cycle import LESSONS_PER_CYCLE

# Открытая сделка старше этого порога считается ПОТЕРЕЙ, а не «в работе»:
# цикл продления укладывается в месяц, всё что висит дольше — уже не решается.
# Порог влияет только на отчёт, домен его не знает.
STUCK_AFTER_DAYS = 30

# Диапазоны циклов для разрезов по преподавателям и направлениям. Поцикловая
# разбивка там бессмысленна: у преподавателя 10–50 учеников, на отдельный цикл
# приходится 0–2 сделки, и «0 из 1» читалось бы как катастрофа. Границы
# выбраны по смыслу: 9-й цикл — конец 36-урочного курса, 18-й — конец второго.
CYCLE_BANDS = [
    (1, 4, '1–4 (1–16 уроков)'),
    (5, 9, '5–9 (17–36, до конца курса)'),
    (10, 18, '10–18 (второй курс)'),
    (19, 9999, '19+ (дальше)'),
]

FONT_NAME = 'Arial'

_NOTES = [
    'Переход считается по ФАКТУ дохода до следующего цикла, а не по отметке «Ушёл»: '
    'таких отметок в базе два десятка на всю историю, ушедших обычно просто перестают вести. '
    'Зависшая сделка (открыта дольше '
    f'{STUCK_AFTER_DAYS} дней) считается потерей.',
    '«В работе» — цикл идёт прямо сейчас, решение не принято. Эти сделки исключены из '
    'знаменателя: считать их потерей значило бы занижать переход тем сильнее, чем больше '
    'у преподавателя активных учеников.',
    'В разрезах по преподавателям и направлениям сумма строк больше школьного итога: сделка '
    'привязана к ученику, а не к группе, поэтому занимающийся у двоих учитывается в обеих '
    'строках. Колонку складывать нельзя. Строка «Архив (импорт истории)» — служебная.',
]


# ---------------------------------------------------------------------------
# Данные
# ---------------------------------------------------------------------------

def _fetch_deals() -> list[dict]:
    """Все сделки продления с исходом, циклом и возрастом."""
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT d.id,
                   d.student_id,
                   s.full_name                                       AS student_name,
                   d.cycle_no,
                   st.kind                                           AS kind,
                   st.label                                          AS stage_label,
                   to_char(d.outcome_at AT TIME ZONE 'Europe/Moscow', 'YYYY-MM')
                                                                     AS outcome_month,
                   (d.outcome_at AT TIME ZONE 'Europe/Moscow')::date  AS outcome_date,
                   COALESCE(d.due_at, d.stage_entered_at::date)      AS reference_date,
                   (d.outcome_at IS NULL)                            AS is_open,
                   (d.outcome_at IS NULL
                    AND COALESCE(d.due_at, d.stage_entered_at::date)
                        < current_date - INTERVAL '{STUCK_AFTER_DAYS} days') AS is_stuck,
                   CASE WHEN d.outcome_at IS NULL
                        THEN current_date - COALESCE(d.due_at, d.stage_entered_at::date)
                   END                                               AS days_open
              FROM renewal_deal d
              JOIN renewal_stage st ON st.id = d.stage_id
              JOIN students s       ON s.id = d.student_id
             ORDER BY d.cycle_no, s.full_name
        """)
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def _student_links() -> tuple[dict[int, list[dict]], dict[int, list[dict]]]:
    """
    student_id → преподаватели и student_id → направления, через ЛЮБЫЕ членства.

    Членство неактивное и группа архивная считаются: ушедший ученик — часть
    истории переходимости, выкинув его, мы показали бы только выживших.
    """
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT m.student_id,
                            t.id AS teacher_id, t.name AS teacher_name, t.is_service,
                            dir.id AS direction_id, dir.name AS direction_name
              FROM group_memberships m
              JOIN groups g       ON g.id = m.group_id
              JOIN teachers t     ON t.id = g.teacher_id
              JOIN directions dir ON dir.id = g.direction_id
        """)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    teachers: dict[int, dict[int, dict]] = defaultdict(dict)
    directions: dict[int, dict[int, dict]] = defaultdict(dict)
    for r in rows:
        teachers[r['student_id']][r['teacher_id']] = {
            'id': r['teacher_id'], 'name': r['teacher_name'], 'is_service': r['is_service'],
        }
        directions[r['student_id']][r['direction_id']] = {
            'id': r['direction_id'], 'name': r['direction_name'], 'is_service': False,
        }
    return (
        {sid: list(v.values()) for sid, v in teachers.items()},
        {sid: list(v.values()) for sid, v in directions.items()},
    )


def _classify(deal: dict) -> str:
    """Исход сделки для воронки: advanced / lost / stuck / active."""
    if deal['kind'] == 'won':
        return 'advanced'
    if deal['kind'] == 'lost':
        return 'lost'
    return 'stuck' if deal['is_stuck'] else 'active'


def _rate(bucket: dict) -> float | None:
    """
    Доля перешедших на следующий цикл.

    Знаменатель — только РЕШЁННЫЕ сделки: перешли + отмеченные уходы + зависшие.
    «В работе» не входят (см. оговорку 2). None, когда решённых нет вовсе:
    0 % и «ещё никто не дошёл» — противоположные вещи.
    """
    decided = bucket['advanced'] + bucket['lost'] + bucket['stuck']
    return bucket['advanced'] / decided if decided else None


def funnel_by_cycle(deals: list[dict]) -> list[dict]:
    """Строка на цикл: сколько дошло и чем закончилось."""
    acc: dict[int, dict] = {}
    for deal in deals:
        bucket = acc.setdefault(deal['cycle_no'], {
            'advanced': 0, 'lost': 0, 'stuck': 0, 'active': 0,
        })
        bucket[_classify(deal)] += 1

    rows = []
    for cycle in sorted(acc):
        bucket = acc[cycle]
        reached = sum(bucket.values())
        rows.append({
            'cycle': cycle,
            'lessons_to': cycle * LESSONS_PER_CYCLE,
            'reached': reached,
            **bucket,
            'rate': _rate(bucket),
        })
    return rows


def _band_of(cycle: int) -> str:
    for low, high, label in CYCLE_BANDS:
        if low <= cycle <= high:
            return label
    return CYCLE_BANDS[-1][2]


def bands_by_entity(deals: list[dict], links: dict[int, list[dict]]) -> list[dict]:
    """Строка на сущность, колонки — диапазоны циклов."""
    acc: dict[int, dict] = {}
    names: dict[int, dict] = {}

    for deal in deals:
        band = _band_of(deal['cycle_no'])
        outcome = _classify(deal)
        for entity in links.get(deal['student_id'], ()):
            row = acc.setdefault(entity['id'], {
                'students': set(),
                'bands': {label: {'advanced': 0, 'lost': 0, 'stuck': 0, 'active': 0}
                          for _lo, _hi, label in CYCLE_BANDS},
            })
            names[entity['id']] = entity
            row['students'].add(deal['student_id'])
            row['bands'][band][outcome] += 1

    out = []
    for entity_id, row in acc.items():
        total_stuck = sum(b['stuck'] for b in row['bands'].values())
        out.append({
            'id': entity_id,
            'name': names[entity_id]['name'],
            'is_service': names[entity_id]['is_service'],
            'students': len(row['students']),
            'stuck': total_stuck,
            'bands': {label: {**counts, 'rate': _rate(counts)}
                      for label, counts in row['bands'].items()},
        })
    # Служебные — в конец, живые — по числу учеников.
    out.sort(key=lambda r: (r['is_service'], -r['students'], r['name']))
    return out


def stuck_deals(deals: list[dict], links_t: dict, links_d: dict) -> list[dict]:
    """
    Рабочий список: с кем разобраться сейчас.

    Сортировка не просто «самые давние сверху». Сначала идут ученики ЖИВЫХ
    преподавателей, и только потом те, кто числится лишь за служебной записью
    «Архив (импорт истории)»: у последних сделки висят по 1300+ дней и, отсортируй
    мы только по возрасту, они заняли бы весь верх списка и сделали бы его
    нерабочим — а разбираться там не с кем, это следы импорта.
    """
    rows = []
    for deal in deals:
        if not deal['is_stuck']:
            continue
        teachers = links_t.get(deal['student_id'], ())
        rows.append({
            'student_name': deal['student_name'],
            'cycle': deal['cycle_no'],
            'stage_label': deal['stage_label'],
            'days_open': deal['days_open'],
            'reference_date': deal['reference_date'],
            # Только служебные преподаватели (или групп нет вовсе) — разбираться не с кем.
            'service_only': bool(teachers) and all(e['is_service'] for e in teachers)
                            or not teachers,
            'teachers': ', '.join(sorted(e['name'] for e in teachers)),
            'directions': ', '.join(sorted(e['name'] for e in links_d.get(deal['student_id'], ()))),
        })
    rows.sort(key=lambda r: (r['service_only'], -(r['days_open'] or 0), r['student_name']))
    return rows


def collect() -> dict:
    """Все данные отчёта. Отдельно от рендера — тестируется без openpyxl."""
    deals = _fetch_deals()
    links_t, links_d = _student_links()
    return {
        'deals': deals,
        'funnel': funnel_by_cycle(deals),
        'teachers': bands_by_entity(deals, links_t),
        'directions': bands_by_entity(deals, links_d),
        'stuck': stuck_deals(deals, links_t, links_d),
        'links_teachers': links_t,
        'links_directions': links_d,
    }


# ---------------------------------------------------------------------------
# Рендер
# ---------------------------------------------------------------------------

def _base_font(bold=False, size=11, color=None, italic=False):
    from openpyxl.styles import Font
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)


def _write_head(ws, ncols: int, title: str) -> int:
    """Заголовок + оговорки. Возвращает строку, с которой идёт таблица."""
    from openpyxl.styles import Alignment

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _base_font(bold=True, size=14)

    row = 2
    for note in _NOTES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=note)
        c.font = _base_font(size=9, italic=True, color='808080')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 30
        row += 1
    return row + 1


def _style_header(ws, row: int, ncols: int) -> None:
    from openpyxl.styles import Alignment, PatternFill

    fill = PatternFill('solid', fgColor='4F59F9')
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = fill
        cell.font = _base_font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 30


_FUNNEL_COLUMNS = [
    ('Цикл', 8), ('Уроков курса', 14), ('Дошли до цикла', 16),
    ('Перешли дальше', 16), ('Ушли (отмечено)', 16), ('Зависли', 11),
    ('В работе', 11), ('Переход', 11), ('Потеряно', 11),
]


def _write_funnel(ws, rows: list[dict]) -> None:
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    ncols = len(_FUNNEL_COLUMNS)
    head = _write_head(ws, ncols, 'Воронка по циклам: на каком шаге теряем детей')

    for ci, (label, width) in enumerate(_FUNNEL_COLUMNS, start=1):
        ws.cell(row=head, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width
    _style_header(ws, head, ncols)

    first = head + 1
    for i, row in enumerate(rows):
        r = first + i
        ws.cell(row=r, column=1, value=row['cycle']).font = _base_font()
        ws.cell(row=r, column=2, value=row['lessons_to']).font = _base_font()
        for ci, key in ((3, 'reached'), (4, 'advanced'), (5, 'lost'),
                        (6, 'stuck'), (7, 'active')):
            ws.cell(row=r, column=ci, value=row[key]).font = _base_font()
        # Формулы, а не посчитанные числа: лист обязан пересчитываться, если
        # кто-то отфильтрует или поправит данные под собой.
        rate = ws.cell(row=r, column=8, value=f'=IFERROR(D{r}/(D{r}+E{r}+F{r}),"")')
        rate.number_format = '0 %'
        rate.font = _base_font()
        lost = ws.cell(row=r, column=9, value=f'=E{r}+F{r}')
        lost.font = _base_font()

    ws.freeze_panes = ws.cell(row=first, column=3)

    last = first + len(rows) - 1
    if len(rows) < 2:
        return

    # Кривая дожития: сколько учеников доходит до каждого цикла. Одна серия —
    # легенда не нужна, заголовок её называет (правило дата-виза).
    survival = LineChart()
    survival.title = 'Сколько учеников доходит до цикла'
    survival.y_axis.title = 'Учеников'
    survival.x_axis.title = 'Цикл (1 цикл = 4 урока)'
    survival.height, survival.width = 8, 24
    survival.legend = None
    survival.add_data(Reference(ws, min_col=3, min_row=head, max_row=last), titles_from_data=True)
    survival.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    survival.series[0].graphicalProperties.line.width = 20000  # ~2pt, тонкая линия
    ws.add_chart(survival, f'A{last + 3}')

    # Второй график отдельно, а НЕ второй осью на первом: две оси с разными
    # шкалами — главная ошибка в графиках, читатель видит пересечения,
    # которых нет.
    rate_chart = LineChart()
    rate_chart.title = 'Доля перешедших на следующий цикл'
    rate_chart.y_axis.title = 'Переход'
    rate_chart.x_axis.title = 'Цикл'
    rate_chart.height, rate_chart.width = 8, 24
    rate_chart.legend = None
    rate_chart.add_data(Reference(ws, min_col=8, min_row=head, max_row=last), titles_from_data=True)
    rate_chart.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    rate_chart.series[0].graphicalProperties.line.width = 20000
    ws.add_chart(rate_chart, f'A{last + 22}')


def _write_bands(ws, title: str, rows: list[dict]) -> None:
    """Строка = сущность, по паре колонок на диапазон циклов."""
    from openpyxl.utils import get_column_letter

    band_labels = [label for _lo, _hi, label in CYCLE_BANDS]
    ncols = 3 + len(band_labels) * 2
    head = _write_head(ws, ncols, title)

    ws.cell(row=head, column=1, value='Название')
    ws.cell(row=head, column=2, value='Учеников')
    ws.cell(row=head, column=3, value='Зависло всего')
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 14
    for bi, label in enumerate(band_labels):
        col = 4 + bi * 2
        ws.cell(row=head, column=col, value=f'{label}\nпереход')
        ws.cell(row=head, column=col + 1, value=f'{label}\nрешено сделок')
        ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions[get_column_letter(col + 1)].width = 14
    _style_header(ws, head, ncols)

    ri = head
    for row in rows:
        ri += 1
        name = row['name']
        if row['is_service']:
            name = f'{name} — служебная, не в итоге'
        cell = ws.cell(row=ri, column=1, value=name)
        cell.font = _base_font(italic=row['is_service'],
                               color='808080' if row['is_service'] else None)
        ws.cell(row=ri, column=2, value=row['students']).font = _base_font()
        ws.cell(row=ri, column=3, value=row['stuck']).font = _base_font()

        for bi, label in enumerate(band_labels):
            counts = row['bands'][label]
            decided = counts['advanced'] + counts['lost'] + counts['stuck']
            col = 4 + bi * 2
            rate_cell = ws.cell(row=ri, column=col,
                                value=None if not decided else counts['advanced'] / decided)
            rate_cell.number_format = '0 %'
            rate_cell.font = _base_font()
            ws.cell(row=ri, column=col + 1, value=decided).font = _base_font()

    ws.freeze_panes = ws.cell(row=head + 1, column=2)


_STUCK_COLUMNS = [
    ('Ученик', 34), ('Цикл', 8), ('Уроков курса', 14), ('Стадия', 22),
    ('Дней висит', 13), ('Опорная дата', 14), ('Преподаватели', 36),
    ('Направления', 36), ('Разбираться', 13),
]


def _write_stuck(ws, rows: list[dict]) -> None:
    from openpyxl.utils import get_column_letter

    ncols = len(_STUCK_COLUMNS)
    head = _write_head(
        ws, ncols,
        'Зависшие сделки: с кем разобраться сейчас (самые давние сверху)')

    for ci, (label, width) in enumerate(_STUCK_COLUMNS, start=1):
        ws.cell(row=head, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width
    _style_header(ws, head, ncols)

    ri = head
    for row in rows:
        ri += 1
        values = [
            row['student_name'], row['cycle'], row['cycle'] * LESSONS_PER_CYCLE,
            row['stage_label'], row['days_open'], row['reference_date'],
            row['teachers'] or '— нет групп —', row['directions'] or '— нет групп —',
            '' if row['service_only'] else 'да',
        ]
        for ci, value in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            # Следы импорта приглушены: они внизу списка и разбираться там не с кем,
            # но выкидывать их нельзя — сумма по листу должна сходиться с воронкой.
            cell.font = _base_font(italic=row['service_only'],
                                   color='808080' if row['service_only'] else None)

    ws.freeze_panes = ws.cell(row=head + 1, column=2)


_DETAIL_COLUMNS = [
    ('Ученик', 34), ('Цикл', 8), ('Стадия', 22), ('Исход', 16),
    ('Дата исхода', 14), ('Опорная дата', 14), ('Дней открыта', 13),
    ('Преподаватели', 36), ('Направления', 36),
]

_OUTCOME_LABELS = {
    'advanced': 'Перешёл дальше', 'lost': 'Ушёл (отмечено)',
    'stuck': 'Зависла', 'active': 'В работе',
}


def _write_detail(ws, deals: list[dict], links_t: dict, links_d: dict) -> None:
    from openpyxl.utils import get_column_letter

    ncols = len(_DETAIL_COLUMNS)
    head = _write_head(ws, ncols, 'Детализация: строка = сделка продления')

    for ci, (label, width) in enumerate(_DETAIL_COLUMNS, start=1):
        ws.cell(row=head, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width
    _style_header(ws, head, ncols)

    ri = head
    for deal in deals:
        ri += 1
        values = [
            deal['student_name'], deal['cycle_no'], deal['stage_label'],
            _OUTCOME_LABELS[_classify(deal)], deal['outcome_date'],
            deal['reference_date'], deal['days_open'],
            ', '.join(sorted(e['name'] for e in links_t.get(deal['student_id'], ())))
            or '— нет групп —',
            ', '.join(sorted(e['name'] for e in links_d.get(deal['student_id'], ())))
            or '— нет групп —',
        ]
        for ci, value in enumerate(values, start=1):
            ws.cell(row=ri, column=ci, value=value).font = _base_font()

    ws.freeze_panes = ws.cell(row=head + 1, column=2)


def render_workbook(data: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = 'Воронка по циклам'
    _write_funnel(ws, data['funnel'])

    _write_bands(wb.create_sheet('Циклы × преподаватели'),
                 'Переход по диапазонам циклов: преподаватели', data['teachers'])
    _write_bands(wb.create_sheet('Циклы × направления'),
                 'Переход по диапазонам циклов: направления', data['directions'])
    _write_stuck(wb.create_sheet('Зависшие сделки'), data['stuck'])
    _write_detail(wb.create_sheet('Детализация'),
                  data['deals'], data['links_teachers'], data['links_directions'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build() -> tuple[bytes, int, str]:
    """(xlsx-байты, число сделок, имя файла). Параметров нет — отчёт общий."""
    from apps.core.utils.dates import msk_now

    data = collect()
    content = render_workbook(data)
    filename = f'retention_{msk_now():%Y-%m-%d}.xlsx'
    return content, len(data['deals']), filename
