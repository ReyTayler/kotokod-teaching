"""
Построитель отчёта «Ученики по преподавателям за месяц».

Строка отчёта — пара «ученик × группа»: ученик, занимающийся в двух группах
одновременно, даёт две строки со своими преподавателями. Преподаватель —
свойство ГРУППЫ (`groups.teacher`), то есть её основной преподаватель на текущий
момент; фактические замены на отдельных занятиях (`lessons.teacher_id`) в этот
отчёт не попадают — так и просили.

Три неочевидных правила подсчёта:

  • Сгорание урока материализовано отдельной записью-уроком с present=true
    (см. apps/extra_lessons/services.py), поэтому фильтра по present мало —
    lesson_type='burned' исключается явно, иначе списанный урок считался бы
    посещённым.
  • Единица — УРОК, не занятие: инвариант half-lesson (45 минут → 0.5). Вес
    берётся с самого занятия, не с группы: у отработок длительность своя.
  • «Уроков у группы за месяц» — только курсовые занятия (COURSE_LESSON_TYPES).
    Доп.урок и сгорание адресные: они принадлежат ученику, а не сетке группы.
    Поэтому у ученика посещений может оказаться больше, чем провела группа, —
    это отработка пропуска, и так и должно быть.
  • Из сетки группы вычитаются неоплачиваемые пропуски ЭТОГО ученика
    (`unpaid_skip`): такие занятия он не посещает и не оплачивает (перевод,
    заморозка), спрашивать с него их нельзя. Поэтому колонка считается на пару
    «ученик × группа», а не на группу: у двух учеников одной группы числа
    законно разойдутся, а у пропустившего весь месяц будет 0.

См. docs/superpowers/specs/2026-09-03-students-by-teacher-report-design.md
"""
from __future__ import annotations

import datetime
import io
from dataclasses import dataclass
from decimal import Decimal

from django.db.models import Case, DecimalField, Sum, Value, When

from apps.core.utils.dates import msk_month_range
from apps.lessons.models import COURSE_LESSON_TYPES, Lesson, LessonAttendance
from apps.memberships.models import GroupMembership

# Сгорание — не посещение (present=true у записи-факта, см. модуль-docstring).
BURNED_LESSON_TYPE = 'burned'
HALF_LESSON_MINUTES = 45

ZERO = Decimal('0')


def _lesson_weight(duration_field: str) -> Sum:
    """Сумма весов уроков: 45 минут → 0.5, иначе 1 (инвариант half-lesson)."""
    return Sum(
        Case(
            When(**{duration_field: HALF_LESSON_MINUTES}, then=Value(Decimal('0.5'))),
            default=Value(Decimal('1')),
            output_field=DecimalField(max_digits=8, decimal_places=1),
        )
    )


@dataclass
class StudentTeacherRow:
    """Строка отчёта: ученик в одной конкретной группе за месяц."""

    group_name: str
    teacher_name: str
    student_name: str
    group_lessons: Decimal
    lessons: Decimal
    direction_name: str


def _group_month_lessons(month_start: datetime.date, month_end: datetime.date) -> dict:
    """Сколько уроков провела каждая группа за месяц: group_id → сумма весов.

    Только курсовые занятия: доп.урок и сгорание — адресные факты ученика, сетку
    группы они не удлиняют (см. модуль-docstring).
    """
    rows = (
        Lesson.objects
        .filter(
            lesson_date__gte=month_start,
            lesson_date__lte=month_end,
            lesson_type__in=COURSE_LESSON_TYPES,
        )
        .values('group_id')
        .annotate(lessons=_lesson_weight('lesson_duration_minutes'))
    )
    return {r['group_id']: r['lessons'] or ZERO for r in rows}


def _unpaid_skip_pairs(month_start: datetime.date, month_end: datetime.date) -> dict:
    """Неоплачиваемые пропуски месяца: (ученик, группа) → сумма весов уроков.

    Эти занятия ученик не посещает и не оплачивает (перевод, заморозка) — из его
    личной нормы они вычитаются, иначе отчёт требовал бы с него уроки, которых
    для него не было.
    """
    rows = (
        LessonAttendance.objects
        .filter(
            unpaid_skip=True,
            lesson__lesson_date__gte=month_start,
            lesson__lesson_date__lte=month_end,
            lesson__lesson_type__in=COURSE_LESSON_TYPES,
        )
        .values('student_id', 'lesson__group_id')
        .annotate(lessons=_lesson_weight('lesson__lesson_duration_minutes'))
    )
    return {
        (r['student_id'], r['lesson__group_id']): r['lessons'] or ZERO
        for r in rows
    }


def _attended_pairs(month_start: datetime.date, month_end: datetime.date) -> dict:
    """Посещения месяца, свёрнутые до (ученик, группа) → сумма весов уроков."""
    rows = (
        LessonAttendance.objects
        .filter(
            present=True,
            lesson__lesson_date__gte=month_start,
            lesson__lesson_date__lte=month_end,
        )
        .exclude(lesson__lesson_type=BURNED_LESSON_TYPE)
        .values(
            'student_id', 'student__full_name',
            'lesson__group_id', 'lesson__group__name',
            'lesson__group__teacher__name', 'lesson__group__direction__name',
        )
        .annotate(lessons=_lesson_weight('lesson__lesson_duration_minutes'))
    )
    return {
        (r['student_id'], r['lesson__group_id']): StudentTeacherRow(
            group_name=r['lesson__group__name'],
            teacher_name=r['lesson__group__teacher__name'],
            student_name=r['student__full_name'],
            group_lessons=ZERO,  # личная норма ученика — ниже, в collect_month
            lessons=r['lessons'] or ZERO,
            direction_name=r['lesson__group__direction__name'],
        )
        for r in rows
    }


def collect_month(month: str) -> list[StudentTeacherRow]:
    """
    Ученики по преподавателям за месяц: список пар «ученик × группа».

    month: 'YYYY-MM'. Пара попадает в отчёт, если членство активно ЛИБО в месяце
    были занятия этой группы с этим учеником. Второе шире первого намеренно:
    ученик, ходивший в июле и ушедший в августе, обязан остаться в июльском
    отчёте — иначе цифры прошлых месяцев менялись бы задним числом. Ученик без
    единой группы строки не даёт: у него нет преподавателя.

    Raises:
        ValueError: month не в формате YYYY-MM / невалидный месяц (1-12).
    """
    start_str, end_str = msk_month_range(f'{month}-01')
    month_start = datetime.date.fromisoformat(start_str)
    month_end = datetime.date.fromisoformat(end_str)

    group_lessons = _group_month_lessons(month_start, month_end)
    unpaid_skips = _unpaid_skip_pairs(month_start, month_end)
    by_pair = _attended_pairs(month_start, month_end)

    # Активные членства — чтобы ученик, не появившийся за месяц ни разу, тоже
    # дал строку (0 уроков), а не молча исчез из отчёта преподавателя.
    memberships = GroupMembership.objects.filter(active=True).values(
        'student_id', 'student__full_name',
        'group_id', 'group__name', 'group__teacher__name', 'group__direction__name',
    )
    for m in memberships:
        key = (m['student_id'], m['group_id'])
        if key in by_pair:
            continue
        by_pair[key] = StudentTeacherRow(
            group_name=m['group__name'],
            teacher_name=m['group__teacher__name'],
            student_name=m['student__full_name'],
            group_lessons=ZERO,
            lessons=ZERO,
            direction_name=m['group__direction__name'],
        )

    # Личная норма ученика: сетка группы минус его неоплачиваемые пропуски.
    # max(..., 0) — страховка от расхождений в данных (пропусков помечено больше,
    # чем занятий в сетке): отрицательная «норма» в отчёте была бы бессмыслицей.
    for (student_id, group_id), row in by_pair.items():
        norm = group_lessons.get(group_id, ZERO) - unpaid_skips.get(
            (student_id, group_id), ZERO)
        row.group_lessons = norm if norm > ZERO else ZERO

    return sorted(
        by_pair.values(),
        key=lambda r: (r.teacher_name, r.group_name, r.student_name),
    )


HEADERS = [
    'Группа', 'Преподаватель', 'Ученик',
    'Уроков у группы за месяц', 'Посещено учеником', 'Направление',
]
_COLUMN_WIDTHS = {1: 26, 2: 26, 3: 32, 4: 18, 5: 18, 6: 24}
# Колонки-числа: Excel должен их суммировать сам, поэтому float + формат «0.#».
_NUMERIC_COLUMNS = (4, 5)


def build_workbook(rows: list[StudentTeacherRow]):
    """Собрать openpyxl.Workbook: плоский лист без строк итогов.

    Плоский он намеренно — служебные строки ломают автофильтр и сводные таблицы,
    которыми этот отчёт и разбирают.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ученики по преподавателям'

    ws.append(HEADERS)

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='EDEDED')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    number_align = Alignment(horizontal='center')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for i, row in enumerate(rows):
        r_idx = 2 + i
        ws.cell(row=r_idx, column=1, value=row.group_name)
        ws.cell(row=r_idx, column=2, value=row.teacher_name)
        ws.cell(row=r_idx, column=3, value=row.student_name)
        # float, а не Decimal/строка: иначе Excel не просуммирует колонку сам.
        ws.cell(row=r_idx, column=4, value=float(row.group_lessons))
        ws.cell(row=r_idx, column=5, value=float(row.lessons))
        ws.cell(row=r_idx, column=6, value=row.direction_name)
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r_idx, column=col)
            cell.border = border
            if col in _NUMERIC_COLUMNS:
                # Кратно 0.5: показываем «4» и «4,5», без хвоста «4,0».
                cell.number_format = '0.#'
                cell.alignment = number_align

    for col_idx, width in _COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_col = get_column_letter(len(HEADERS))
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = f'A1:{last_col}{max(len(rows) + 1, 2)}'
    return wb


def render_bytes(rows: list[StudentTeacherRow]) -> bytes:
    """Отчёт как xlsx-байты (для раздела «Отчёты»)."""
    wb = build_workbook(rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build(month: str) -> tuple[bytes, int, str]:
    """(xlsx-байты, число строк ученик×группа, имя файла). month — 'YYYY-MM'."""
    rows = collect_month(month)  # ValueError при кривом месяце → services пометит failure
    content = render_bytes(rows)
    filename = f'students_by_teacher_{month}.xlsx'
    return content, len(rows), filename
