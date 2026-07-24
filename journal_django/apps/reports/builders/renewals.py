"""
Построитель «Отчёта по продлениям» за выбранный месяц.

Разделён на чистые функции для тестируемости без Celery/HTTP:
  • collect_rows(year, month)      — данные (SQL по renewal_deal/stage);
  • render_workbook(rows, y, m)    — xlsx-байты (openpyxl).

Дата стадии берётся из САМОЙ сделки, а не из лога renewal_activity: у активности
created_at = момент записи строки, и после массового «rebuild-renewals» он у всех
сделок равен дате пересбора (отсюда была ложная «20.07» у всех). Реальную дату
попадания в стадию хранит сама сделка — та же, что показывает карточка:
  • закрытая сделка → outcome_at («Продлён 28.08.2025» в дровере);
  • открытая «Ждём продление» → due_at (день 4-го урока цикла);
  • прочая открытая → stage_entered_at (когда встала на текущую стадию).
`stage_date = COALESCE(outcome_at, due_at, stage_entered_at)`. Сделка попадает в
отчёт за месяц, если её stage_date — внутри месяца. Строка = сделка (у ученика
может быть несколько циклов → несколько строк).

Точность открытых сделок: бэкфилл `renewals.backfill_open_dates` (раздел Синхро)
восстанавливает реальную stage_entered_at открытых авто-сделок из посещаемости,
не перетирая CRM-данные — после него дата верна и без due_at-фолбэка.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from django.db import connection
from django.utils import timezone

MONTHS_RU = [
    '', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
    'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь',
]


def month_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    """[начало месяца, начало следующего месяца) как tz-aware datetime (Europe/Moscow)."""
    tz = timezone.get_current_timezone()
    start = datetime(year, month, 1, tzinfo=tz)
    nxt = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    end = datetime(nxt.year, nxt.month, nxt.day, tzinfo=tz)
    return start, end


def collect_rows(year: int, month: int) -> list[dict]:
    """
    Строки отчёта: одна на КАЖДУЮ сделку, чья дата стадии (outcome_at для закрытой,
    иначе stage_entered_at) попадает в выбранный месяц. Сортировка по ФИО, затем
    по номеру цикла.
    """
    start, end = month_bounds(year, month)
    sql = """
        SELECT s.full_name                         AS student_name,
               a.full_name                         AS assignee_name,
               st.label                            AS stage_label,
               (d.outcome_at IS NOT NULL)          AS is_closed,
               COALESCE(d.outcome_at, d.due_at::timestamptz, d.stage_entered_at) AS stage_date,
               d.cycle_no                          AS cycle_no
        FROM renewal_deal d
        JOIN students s       ON s.id = d.student_id
        JOIN renewal_stage st ON st.id = d.stage_id
        LEFT JOIN accounts a  ON a.id = d.assignee_id
        WHERE COALESCE(d.outcome_at, d.due_at::timestamptz, d.stage_entered_at) >= %(start)s
          AND COALESCE(d.outcome_at, d.due_at::timestamptz, d.stage_entered_at) <  %(end)s
        ORDER BY s.full_name, d.cycle_no
    """
    with connection.cursor() as cur:
        cur.execute(sql, {'start': start, 'end': end})
        cols = [c[0] for c in cur.description]
        raw = [dict(zip(cols, r)) for r in cur.fetchall()]

    rows: list[dict] = []
    for r in raw:
        closed = r['is_closed']
        stage_date = r['stage_date']
        rows.append({
            'student_name': r['student_name'],
            'assignee_name': r['assignee_name'] or '—',
            'active_stage': '' if closed else r['stage_label'],
            'closed_stage': r['stage_label'] if closed else '',
            'entered_at': timezone.localtime(stage_date).date() if stage_date else None,
            'cycle_no': r['cycle_no'],
        })
    return rows


COLUMNS = [
    ('ФИО ученика', 'student_name', 34),
    ('Цикл сделки', 'cycle_no', 12),
    ('Ответственный за сделку', 'assignee_name', 28),
    ('Текущая стадия активной сделки', 'active_stage', 30),
    ('Стадия закрытой сделки', 'closed_stage', 26),
    ('Дата переноса в эту стадию', 'entered_at', 22),
]


def render_workbook(rows: list[dict], year: int, month: int) -> bytes:
    """Собрать xlsx (openpyxl) из строк collect_rows. Возвращает байты файла."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Продления'

    title = f'Отчёт по продлениям — {MONTHS_RU[month]} {year}'
    ncols = len(COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    header_fill = PatternFill('solid', fgColor='4F59F9')
    header_font = Font(bold=True, color='FFFFFF')
    for ci, (label, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 30

    for ri, row in enumerate(rows, start=3):
        for ci, (_label, key, _w) in enumerate(COLUMNS, start=1):
            value = row.get(key)
            cell = ws.cell(row=ri, column=ci, value=value)
            if key == 'entered_at' and value is not None:
                cell.number_format = 'DD.MM.YYYY'
                cell.alignment = Alignment(horizontal='center')
            elif key == 'cycle_no':
                cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(ncols)}{max(2, len(rows) + 2)}'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build(year: int, month: int) -> tuple[bytes, int, str]:
    """Полная сборка: (xlsx-байты, число строк, имя файла)."""
    rows = collect_rows(year, month)
    content = render_workbook(rows, year, month)
    filename = f'renewals_{year}-{month:02d}.xlsx'
    return content, len(rows), filename
