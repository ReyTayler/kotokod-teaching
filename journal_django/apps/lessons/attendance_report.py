"""
Сборка данных отчёта по посещаемости за месяц + запись в Excel.

Строка отчёта — пара (ученик × группа); в листе она занимает ДВЕ строки:
верхняя — даты уроков, нижняя — статусы. Колонки «Урок N» — порядковые уроки
самого ученика внутри месяца, а не календарные даты: у разных учеников в одной
колонке стоят разные даты.

Ключевое правило: отработка (доп.урок) и сгорание НЕ создают лишнюю колонку —
они «ложатся» на ту позицию, где был пропуск, меняя её статус. Иначе отчёт
читается так, будто ученика не было на уроке 2 и он был на уроке 3, хотя на
деле он отработал именно за урок 2.

В БД и доп.урок, и сгорание материализованы как отдельные записи-уроки
(lesson_type 'extra'/'burned' + строка lesson_attendance, см.
apps/extra_lessons/services.py), связанные с пропуском через AbsenceResolution.
Поэтому схлопывание — это подавление колонки факта при живом пропуске в ТОМ ЖЕ
месяце. Если пропуск в другом месяце (в базе это обычное дело: пропуск 14.06 →
отработка 19.07), схлопывать не на что — факт показывается своей колонкой с
пояснением, за какой день отработка. Так каждое событие попадает ровно в один
месячный отчёт и никогда не задваивается между соседними.

См. docs/superpowers/specs/2026-07-27-attendance-monthly-report-design.md
"""
from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from pathlib import Path

from django.db.models import Q

from apps.core.utils.dates import msk_month_range
from apps.extra_lessons.models import BURNED as RESOLUTION_BURNED
from apps.extra_lessons.models import EXTRA as RESOLUTION_EXTRA
from apps.extra_lessons.models import MAKEUP_DONE
from apps.lessons.models import LessonAttendance
from apps.memberships.models import GroupMembership
from apps.students.models import Student

# Категории ячеек — по ним считаются итоговые колонки. Отделены от подписи:
# подпись бывает динамической («Отработка за 14.06.2026»), а счётчику нужна
# устойчивая категория.
KIND_PRESENT = 'present'
KIND_ABSENT = 'absent'
KIND_MADE_UP = 'made_up'
KIND_BURNED = 'burned'

PRESENT = 'Был'
ABSENT = 'Не был'
BURNED = 'Сгорел'
MADE_UP = 'Отработал'
EXTRA_OVER_COURSE = 'Доп.урок'
EMPTY = '-'
# Ученик без единой группы (ни уроков за месяц, ни активных членств).
NO_GROUP = '—'


def _fmt(d: datetime.date) -> str:
    return d.strftime('%d.%m.%Y')


@dataclass
class AttendanceCell:
    """Один урок ученика: дата проведения, подпись статуса и его категория."""

    date: datetime.date
    status: str
    kind: str


@dataclass
class AttendanceRow:
    """Пара «ученик × группа» — единица отчёта (две строки на листе)."""

    student_id: int
    full_name: str
    group_label: str
    cells: list[AttendanceCell] = field(default_factory=list)

    def _count(self, kind: str) -> int:
        return sum(1 for c in self.cells if c.kind == kind)

    @property
    def present_count(self) -> int:
        return self._count(KIND_PRESENT)

    @property
    def absent_count(self) -> int:
        return self._count(KIND_ABSENT)

    @property
    def made_up_count(self) -> int:
        return self._count(KIND_MADE_UP)


def _resolution_maps(month_start: datetime.date, month_end: datetime.date):
    """
    Резолюции пропусков, касающиеся месяца, в двух разрезах.

    Возвращает (by_fact, by_miss), обе с ключом (lesson_id, student_id):
      by_fact — резолюция, ФАКТОМ которой является этот урок (доп.урок/сгорание);
      by_miss — резолюция ПРОПУСКА на этом уроке.

    Берём резолюции, у которых в месяц попадает хотя бы одна сторона (пропуск
    или факт) — этого достаточно, чтобы принять решение по каждой ячейке месяца.
    """
    from apps.extra_lessons.models import AbsenceResolution

    rows = AbsenceResolution.objects.filter(
        Q(fact_lesson__lesson_date__gte=month_start, fact_lesson__lesson_date__lte=month_end)
        | Q(missed_lesson__lesson_date__gte=month_start, missed_lesson__lesson_date__lte=month_end)
    ).values(
        'student_id', 'kind', 'status',
        'missed_lesson_id', 'missed_lesson__lesson_date',
        'fact_lesson_id', 'fact_lesson__lesson_date',
    )

    by_fact: dict[tuple[int, int], dict] = {}
    by_miss: dict[tuple[int, int], dict] = {}
    for r in rows:
        if r['fact_lesson_id'] is not None:
            by_fact[(r['fact_lesson_id'], r['student_id'])] = r
        if r['missed_lesson_id'] is not None:
            by_miss[(r['missed_lesson_id'], r['student_id'])] = r
    return by_fact, by_miss


def _cell_for(
    att: dict,
    by_fact: dict,
    by_miss: dict,
    month_start: datetime.date,
    month_end: datetime.date,
) -> AttendanceCell | None:
    """
    Ячейка для одной записи посещаемости, либо None — если колонку показывать
    не нужно (факт отработки/сгорания схлопнулся в ячейку своего пропуска).
    """
    key = (att['lesson_id'], att['student_id'])
    date = att['lesson__lesson_date']
    lesson_type = att['lesson__lesson_type']

    def in_month(d: datetime.date | None) -> bool:
        return d is not None and month_start <= d <= month_end

    # 1. Этот урок — факт резолюции (доп.урок или сгорание).
    fact = by_fact.get(key)
    if fact is not None:
        missed_date = fact['missed_lesson__lesson_date']
        if in_month(missed_date):
            # Пропуск виден в этом же отчёте — факт ложится на его ячейку,
            # своей колонки не занимает (см. п.3 ниже).
            return None
        if fact['kind'] == RESOLUTION_EXTRA:
            # Доп.урок СВЕРХ курса: пропуска нет вовсе, ложиться не на что.
            return AttendanceCell(date=date, status=EXTRA_OVER_COURSE, kind=KIND_PRESENT)
        # Пропуск в другом месяце: показываем факт своей датой и поясняем, за какой
        # день он закрыт, — иначе занятие/списание этого месяца пропало бы из отчёта.
        if lesson_type == 'burned':
            return AttendanceCell(date=date, status=f'{BURNED} (за {_fmt(missed_date)})',
                                  kind=KIND_BURNED)
        return AttendanceCell(date=date, status=f'Отработка за {_fmt(missed_date)}',
                              kind=KIND_MADE_UP)

    # 2. Сгорание без резолюции (страховка на исторические записи).
    if lesson_type == 'burned':
        return AttendanceCell(date=date, status=BURNED, kind=KIND_BURNED)

    if att['present']:
        return AttendanceCell(date=date, status=PRESENT, kind=KIND_PRESENT)

    # 3. Пропуск: если он закрыт в ЭТОМ же месяце — показываем исход прямо здесь,
    # на месте пропущенного урока. Закрытие в другом месяце сюда не тянем: оно
    # покажется в своём отчёте (п.1), иначе одно событие попало бы в два месяца.
    miss = by_miss.get(key)
    if miss is not None and in_month(miss['fact_lesson__lesson_date']):
        if miss['status'] == MAKEUP_DONE:
            return AttendanceCell(date=date, status=MADE_UP, kind=KIND_MADE_UP)
        if miss['status'] == RESOLUTION_BURNED:
            return AttendanceCell(date=date, status=BURNED, kind=KIND_BURNED)

    return AttendanceCell(date=date, status=ABSENT, kind=KIND_ABSENT)


def collect_month(month: str) -> list[AttendanceRow]:
    """
    Посещаемость ВСЕХ учеников системы за указанный месяц.

    month: 'YYYY-MM'. В отчёт попадают уроки с lesson_date внутри месяца
    включительно. Ученик присутствует в результате всегда — даже без единого
    урока (строка сплошных прочерков).

    Raises:
        ValueError: month не в формате YYYY-MM / невалидный месяц (1-12).
    """
    start_str, end_str = msk_month_range(f'{month}-01')
    month_start = datetime.date.fromisoformat(start_str)
    month_end = datetime.date.fromisoformat(end_str)

    by_fact, by_miss = _resolution_maps(month_start, month_end)

    # 1. Посещения за месяц. Порядок задаётся здесь — дальше только раскладка,
    # без пересортировки: ФИО → группа → дата → номер урока → id (стабильный
    # тай-брейк, если в один день у группы два урока с одним номером).
    attendance = (
        LessonAttendance.objects
        .filter(lesson__lesson_date__gte=month_start, lesson__lesson_date__lte=month_end)
        .order_by(
            'student__full_name', 'lesson__group__name',
            'lesson__lesson_date', 'lesson__lesson_number', 'lesson_id',
        )
        .values(
            'lesson_id', 'student_id', 'lesson__group_id', 'lesson__group__name',
            'lesson__lesson_date', 'lesson__lesson_type', 'present',
        )
    )
    cells_by_pair: dict[tuple[int, int], list[AttendanceCell]] = {}
    group_names: dict[int, str] = {}
    for att in attendance:
        cell = _cell_for(att, by_fact, by_miss, month_start, month_end)
        if cell is None:
            continue
        gid = att['lesson__group_id']
        group_names[gid] = att['lesson__group__name']
        cells_by_pair.setdefault((att['student_id'], gid), []).append(cell)

    # 2. Активные членства — чтобы группа, которая в этом месяце не занималась,
    # тоже дала строку (сплошные прочерки), а не молча исчезла из отчёта.
    memberships: dict[int, set[int]] = {}
    for m in GroupMembership.objects.filter(active=True).values(
        'student_id', 'group_id', 'group__name'
    ):
        group_names.setdefault(m['group_id'], m['group__name'])
        memberships.setdefault(m['student_id'], set()).add(m['group_id'])

    # 3. Все ученики базы — требование руководства: в таблице обязаны быть все.
    rows: list[AttendanceRow] = []
    for s in Student.objects.order_by('full_name').values('id', 'full_name'):
        sid = s['id']
        group_ids = {gid for (st_id, gid) in cells_by_pair if st_id == sid}
        group_ids |= memberships.get(sid, set())
        if not group_ids:
            rows.append(AttendanceRow(student_id=sid, full_name=s['full_name'],
                                      group_label=NO_GROUP))
            continue
        for gid in sorted(group_ids, key=lambda g: group_names.get(g, '')):
            rows.append(AttendanceRow(
                student_id=sid,
                full_name=s['full_name'],
                group_label=group_names.get(gid, NO_GROUP),
                cells=cells_by_pair.get((sid, gid), []),
            ))
    return rows


def build_workbook(rows: list[AttendanceRow]):
    """Собрать openpyxl.Workbook отчёта (пара ученик×группа = две строки листа).

    Общее ядро для write_xlsx (файл на диск, CLI-команда) и render_bytes
    (байты для раздела «Отчёты»)."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    max_lessons = max((len(r.cells) for r in rows), default=0)

    name_col, group_col = 1, 2
    lesson_base = 3                                  # первая колонка «Урок 1»
    present_col = lesson_base + max_lessons
    absent_col = present_col + 1
    made_up_col = absent_col + 1
    last_col = made_up_col

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Посещаемость'

    headers = ['ФИО ученика', 'Группа']
    headers += [f'Урок {i}' for i in range(1, max_lessons + 1)]
    headers += ['Итого «Был»', 'Итого «Не был»', 'Итого «Отработано»']
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='EDEDED')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    date_fmt = 'DD.MM.YYYY'
    total_cols = (present_col, absent_col, made_up_col)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for i, row in enumerate(rows):
        top = 2 + 2 * i          # строка дат
        bottom = top + 1         # строка статусов

        ws.cell(row=top, column=name_col, value=row.full_name)
        ws.cell(row=top, column=group_col, value=row.group_label)
        ws.cell(row=top, column=present_col, value=row.present_count)
        ws.cell(row=top, column=absent_col, value=row.absent_count)
        ws.cell(row=top, column=made_up_col, value=row.made_up_count)
        # ФИО / группа / итоги — общие для пары строк, объединяем по вертикали.
        for col in (name_col, group_col, *total_cols):
            ws.merge_cells(start_row=top, start_column=col, end_row=bottom, end_column=col)
            ws.cell(row=top, column=col).alignment = Alignment(
                horizontal='center' if col in total_cols else 'left',
                vertical='center',
            )

        for j in range(max_lessons):
            col = lesson_base + j
            date_cell = ws.cell(row=top, column=col)
            status_cell = ws.cell(row=bottom, column=col)
            if j < len(row.cells):
                date_cell.value = row.cells[j].date
                date_cell.number_format = date_fmt
                status_cell.value = row.cells[j].status
            else:
                date_cell.value = EMPTY
                status_cell.value = EMPTY
            date_cell.alignment = center
            status_cell.alignment = center

        for r_idx in (top, bottom):
            for col in range(1, last_col + 1):
                ws.cell(row=r_idx, column=col).border = border

    widths = {name_col: 32, group_col: 26,
              present_col: 13, absent_col: 15, made_up_col: 19}
    for j in range(max_lessons):
        # Шире обычного: подпись «Отработка за 14.06.2026» длиннее «Был».
        widths[lesson_base + j] = 20
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Шапка + ФИО/группа остаются видимыми при прокрутке вправо и вниз.
    ws.freeze_panes = ws.cell(row=2, column=lesson_base)
    return wb


def write_xlsx(rows: list[AttendanceRow], path: str | Path) -> None:
    """Пишет отчёт в файл на диск (CLI-команда export_attendance_report)."""
    wb = build_workbook(rows)
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def render_bytes(rows: list[AttendanceRow]) -> bytes:
    """Отчёт как xlsx-байты (для раздела «Отчёты»)."""
    import io
    wb = build_workbook(rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
